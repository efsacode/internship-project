import sys
import pandas as pd
import openpyxl
import os
from PyQt5.QtWidgets import QDialogButtonBox
from PyQt5.QtWidgets import QDialog
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QTextEdit
from PyQt5.QtCore import Qt, QSize
from PyQt5.QtGui import QIcon, QFont
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QLineEdit, QMessageBox, QComboBox, QStackedWidget, QInputDialog, QSpinBox, QDoubleSpinBox, QScrollArea, QCheckBox, QRadioButton, QFileDialog, QSizePolicy,QGroupBox,QFormLayout,QTabWidget, QTableWidgetItem, QListWidget, QAbstractItemView, QListWidgetItem
 
class OpeningPage(QtWidgets.QWidget):
   
    def __init__(self, stacked_widget):
        super().__init__()
        self.stacked_widget = stacked_widget
        self.setupUi()
       

    def setupUi(self):
        self.setObjectName("OpeningPage")
        self.setMinimumSize(412, 286)

        
        font = QtGui.QFont()
        font.setPointSize(30)

       
        main_layout = QtWidgets.QHBoxLayout(self)
        main_layout.setContentsMargins(50, 50, 50, 50)
        main_layout.setSpacing(20)
        main_layout.setAlignment(QtCore.Qt.AlignCenter)  #ortalamak icin

        
        main_layout.addStretch()

        
        user_layout = QtWidgets.QHBoxLayout()
        user_layout.setContentsMargins(0, 0, 0, 0)
        user_layout.setSpacing(20)
        user_layout.setAlignment(QtCore.Qt.AlignCenter)  

        self.user_label = QtWidgets.QLabel("User Sign In", self)
        self.user_label.setFont(font)
        user_layout.addWidget(self.user_label)

        self.user_pushButton = QtWidgets.QPushButton("Sign In", self)
        
        self.user_pushButton.setObjectName("user_pushButton")
        self.user_pushButton.clicked.connect(self.show_user_signin_page)
        self.user_pushButton.setFont(QFont("Arial",20))
        user_layout.addWidget(self.user_pushButton)

        main_layout.addLayout(user_layout)

        
        main_layout.addStretch()

       
        admin_layout = QtWidgets.QHBoxLayout()
        admin_layout.setContentsMargins(0, 0, 0, 0)
        admin_layout.setSpacing(20)
        admin_layout.setAlignment(QtCore.Qt.AlignCenter)  

        self.admin_label = QtWidgets.QLabel("Admin Sign In", self)
        self.admin_label.setFont(font)
        admin_layout.addWidget(self.admin_label)

        self.admin_pushButton = QtWidgets.QPushButton("Sign In", self)
        self.admin_pushButton.setObjectName("admin_pushButton")
        self.admin_pushButton.clicked.connect(self.show_admin_signin_page)
        self.admin_pushButton.setFont(QFont("Arial",20))
        admin_layout.addWidget(self.admin_pushButton)

        main_layout.addLayout(admin_layout)

        
        main_layout.addStretch()

        self.setLayout(main_layout)

        
        self.setSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)


    def show_user_signin_page(self):
        self.stacked_widget.setCurrentIndex(1)
    def show_admin_signin_page(self):
        self.stacked_widget.setCurrentIndex(10)

        
class UserSignInPage(QtWidgets.QWidget):
    
    def __init__(self, stacked_widget):
        super().__init__()
        self.stacked_widget = stacked_widget
        self.setupUi()

        self.user_data = pd.read_excel('kullanici_bilgileri_3.xlsx', sheet_name='Sheet1', engine='openpyxl')

    def setupUi(self):
        self.setObjectName("UserSignInPage")
        self.resize(302, 311)

        font4 = QtGui.QFont()
        font4.setPointSize(12) 

        self.label = QLabel(self)
        self.label.setGeometry(QtCore.QRect(850, 200, 150, 400))
        font = QtGui.QFont()
        font.setPointSize(20)
        self.label.setFont(font)
        self.label.setObjectName("label")

        self.username_input = QLineEdit(self)
        self.username_input.setGeometry(QtCore.QRect(850, 430, 200, 30))
        self.username_input.setObjectName("username_input")
        self.username_input.setFont(QtGui.QFont("Arial", 12))

        self.label_2 = QLabel(self)
        self.label_2.setGeometry(QtCore.QRect(850, 500, 300, 30))
        font3 = QtGui.QFont()
        font3.setPointSize(20)
        self.label_2.setFont(font3)
        self.label_2.setObjectName("label_2")

        self.password_input = QLineEdit(self)
        self.password_input.setGeometry(QtCore.QRect(850, 545, 200, 30))
        self.password_input.setObjectName("password_input")
        self.password_input.setEchoMode(QLineEdit.Password)  
        self.password_input.setFont(QtGui.QFont("Arial", 12))
        
        self.label_3 = QLabel(self)
        self.label_3.setGeometry(QtCore.QRect(850, 300, 300, 50))
        font2 = QtGui.QFont()
        font2.setPointSize(25)
        self.label_3.setFont(font2)
        self.label_3.setObjectName("label_3")

        self.button_container = QWidget(self)
        self.button_container.setGeometry(QtCore.QRect(840, 600, 220, 100))

        self.button_layout = QHBoxLayout(self.button_container)

        self.signButton = QPushButton("Sign In", self.button_container)
        self.signButton.setObjectName("signButton")
        self.signButton.setFixedSize(100, 40)
        self.signButton.setFont(font4)
        self.button_layout.addWidget(self.signButton)
        self.signButton.clicked.connect(self.signin_clicked)

        self.back_button = QPushButton("Go Back", self.button_container)
        self.back_button.setObjectName("back_button")
        self.back_button.setFixedSize(100, 40)
        self.back_button.setFont(font4)
        self.back_button.clicked.connect(self.go_back)
        self.button_layout.addWidget(self.back_button)

        self.retranslateUi()

    def retranslateUi(self):
        _translate = QtCore.QCoreApplication.translate
        self.label.setText(_translate("UserSignInPage", "Username"))
        self.label_2.setText(_translate("UserSignInPage", "Password"))
        self.label_3.setText(_translate("UserSignInPage", "User Sign In"))
        #self.checkBox.setText(_translate("UserSignInPage", "Remember me"))
        self.signButton.setText(_translate("UserSignInPage", "Sign In"))

    def signin_clicked(self):
        username_input = self.username_input.text()
        password_input = self.password_input.text()

       
        def is_valid_username(username):
            try:
                int(username)  
                return self.user_data['Username'].apply(lambda x: str(x)).eq(username).any()
            except ValueError:
                return self.user_data['Username'].eq(username).any()

        
        def is_valid_password(password):
            try:
                int(password)  
                return self.user_data['Password'].apply(lambda x: str(x)).eq(password).any()
            except ValueError:
                return self.user_data['Password'].eq(password).any()

       
        username_valid = is_valid_username(username_input) 
        password_valid = is_valid_password(password_input)

        if username_valid and password_valid:
            QMessageBox.information(self, 'Success', 'Sign in successful!')
            self.parent().show_user_decision_page()
            
            self.password_input.clear() #remember me eklenip degistirilebilir
            self.username_input.clear()
        
        else:
            QMessageBox.warning(self, 'Error', 'Invalid username or password.')

    def go_back(self):
        self.stacked_widget.setCurrentIndex(0)


class AdminSignInPage(QtWidgets.QWidget):
    def __init__(self, stacked_widget):
        super().__init__()
        self.stacked_widget = stacked_widget
        self.setupUi()
        self.user_data = pd.read_excel('kullanici_bilgileri_3.xlsx', sheet_name='Sheet1', engine='openpyxl') #kullanici bilgileri icin bu excel cekiliyor

    def setupUi(self):
        self.setObjectName("AdminSignInPage")
        self.resize(302, 311)

        font4 = QtGui.QFont()
        font4.setPointSize(12) #butonlar icin
        
        self.label = QLabel(self)
        self.label.setGeometry(QtCore.QRect(850, 200, 150,400))
        font = QtGui.QFont()
        font.setPointSize(20)
        self.label.setFont(font)
        self.label.setObjectName("label")
        
        self.username_input = QLineEdit(self)
        self.username_input.setGeometry(QtCore.QRect(850, 430, 200, 30))
        self.username_input.setFont(QtGui.QFont("Arial", 12))
        self.username_input.setObjectName("username_input")
        
        self.label_2 = QLabel(self)
        self.label_2.setGeometry(QtCore.QRect(850, 500, 300, 30))
        font3 = QtGui.QFont()
        font3.setPointSize(20)
        self.label_2.setFont(font3)
        self.label_2.setObjectName("label_2")
        
        self.password_input = QLineEdit(self)
        self.password_input.setGeometry(QtCore.QRect(850, 545, 200, 30))
        self.password_input.setObjectName("password_input")
        self.password_input.setFont(QtGui.QFont("Arial", 12))
        self.password_input.setEchoMode(QLineEdit.Password)  # Şifre giriş modu
        
        self.label_3 = QLabel(self)
        self.label_3.setGeometry(QtCore.QRect(850, 300, 300, 50))
        font2 = QtGui.QFont()
        font2.setPointSize(25)
        self.label_3.setFont(font2)
        self.label_3.setObjectName("label_3")
        
        #self.checkBox = QCheckBox(self)
        #self.checkBox.setGeometry(QtCore.QRect(850, 550, 91, 17))
        #self.checkBox.setObjectName("checkBox")
       
        self.button_container = QWidget(self)
        self.button_container.setGeometry(QtCore.QRect(840, 600, 220, 100))
        
        self.button_layout = QHBoxLayout(self.button_container)
        
        self.signButton = QPushButton("Sign In", self.button_container)
        self.signButton.setObjectName("signButton")
        self.signButton.setFixedSize(100, 40)
         # yazi boyutu 
        self.signButton.setFont(font4)
        self.button_layout.addWidget(self.signButton)
        
        self.back_button = QPushButton("Go Back", self.button_container)
        self.back_button.setObjectName("back_button")
        self.back_button.setFixedSize(100, 40)
        self.back_button.setFont(font4)
        self.back_button.clicked.connect(self.go_back)
        self.button_layout.addWidget(self.back_button)
        self.retranslateUi()

        
        self.signButton.clicked.connect(self.signin_clicked)

    def retranslateUi(self):
        _translate = QtCore.QCoreApplication.translate
        self.label.setText(_translate("AdminSignInPage", "Username"))
        self.label_2.setText(_translate("AdminSignInPage", "Password"))
        self.label_3.setText(_translate("AdminSignInPage", "Admin Sign In"))
        #self.checkBox.setText(_translate("AdminSignInPage", "Remember me")) #belki eklenebiliir diye eklemistim 
        self.signButton.setText(_translate("AdminSignInPage", "Sign In"))

    def go_back(self):
        self.stacked_widget.setCurrentIndex(0)
    
    def signin_clicked(self):
        username_input = self.username_input.text()
        password_input = self.password_input.text()

        role = self.authenticate(username_input, password_input)

        if role == 'Admin':
          
            def is_valid_username(username):
                try:
                    int(username)  
                    return self.user_data['Username'].apply(lambda x: str(x)).eq(username).any()
                except ValueError:
                    return self.user_data['Username'].eq(username).any()

           
            def is_valid_password(password):
                try:
                    int(password)  # Check if password is an integer
                    return self.user_data['Password'].apply(lambda x: str(x)).eq(password).any()
                except ValueError:
                    return self.user_data['Password'].eq(password).any()

            
            username_valid = is_valid_username(username_input)
            password_valid = is_valid_password(password_input)

            if username_valid and password_valid:
                QMessageBox.information(self, 'Success', 'Sign in successful!')
                self.parent().show_admin_decision_page()
                self.password_input.clear() #remember me eklenip degistirilebilir
                self.username_input.clear()
                
            else:
                QMessageBox.warning(self, 'Error', 'Invalid username or password.')
        else:
            QMessageBox.warning(self, 'Error', 'You cannot access here.')
            

    def authenticate(self, username, password):
        for user in self.user_data.itertuples(index=False, name=None):
            if str(user[0]) == username and str(user[1]) == password:  # username is first column, password is second column
                return user[2]  # role donduruldu
        return None

class UserDecisionPage(QtWidgets.QWidget):
    def __init__(self, stacked_widget):
        super().__init__()
        self.stacked_widget = stacked_widget
        self.setupUi()

    def setupUi(self):
        self.setObjectName("UserDecisionPage")
        self.resize(558, 308)
        self.setSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        
        main_layout = QtWidgets.QHBoxLayout(self)

        
        font = QtGui.QFont()
        font.setPointSize(20)

        
        hbox1 = QtWidgets.QVBoxLayout()
        self.label_1 = QtWidgets.QLabel("EDIT EXISTED DATA", self)
        self.label_1.setFont(font)
        hbox1.addWidget(self.label_1)
        
        self.pushButton_1 = QtWidgets.QPushButton("Edit", self)
        self.pushButton_1.setObjectName("edit_button")
        self.pushButton_1.clicked.connect(self.go_to_edit_page)
        self.pushButton_1.setFont(QtGui.QFont("Arial", 20))
        hbox1.addWidget(self.pushButton_1)
        hbox1.setAlignment(QtCore.Qt.AlignCenter)
        hbox1.addSpacing(5) 
        main_layout.addLayout(hbox1)

       
        hbox2 = QtWidgets.QVBoxLayout()
        self.label_2 = QtWidgets.QLabel("CREATE A DATA", self)
        self.label_2.setFont(font)
        hbox2.addWidget(self.label_2)
        
        self.pushButton_2 = QtWidgets.QPushButton("Create", self)
        self.pushButton_2.setObjectName("create_button")
        self.pushButton_2.clicked.connect(self.go_to_baslangic_page)
        self.pushButton_2.setFont(QtGui.QFont("Arial", 20))
        hbox2.setAlignment(QtCore.Qt.AlignCenter)
        hbox2.addWidget(self.pushButton_2)
        
        main_layout.addLayout(hbox2)

        
        hbox3 = QtWidgets.QVBoxLayout()
        self.label_3 = QtWidgets.QLabel("PREVIEW EXISTED DATA", self)
        self.label_3.setFont(font)
        hbox3.addWidget(self.label_3)
        
        self.pushButton_3 = QtWidgets.QPushButton("Preview", self)
        self.pushButton_3.setObjectName("preview_button")
        self.pushButton_3.clicked.connect(self.go_to_prev_page)
        self.pushButton_3.setFont(QtGui.QFont("Arial", 20))
        hbox3.addWidget(self.pushButton_3)
        hbox3.setAlignment(QtCore.Qt.AlignCenter)
        main_layout.addLayout(hbox3)

        
        self.back_button = QPushButton('⬅️', self)
        self.back_button.clicked.connect(self.go_back)
        self.back_button.setFont(QtGui.QFont("Arial", 16))
        

    def retranslateUi(self):
        
        pass

    def go_to_baslangic_page(self):
        self.stacked_widget.setCurrentIndex(3)
    
    def go_to_prev_page(self):
        self.stacked_widget.setCurrentIndex(8)
    
    def go_to_edit_page(self):
        self.stacked_widget.setCurrentIndex(9)
    
    def go_back(self):
        self.stacked_widget.setCurrentIndex(1) 

class AdminDecisionPage(QtWidgets.QWidget):
    
    def __init__(self, stacked_widget):
        super().__init__()
        self.stacked_widget = stacked_widget
        self.setupUi()

    def setupUi(self):
        self.setObjectName("AdminDecisionPage")
        self.resize(558, 308)
        self.setSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        
        main_layout = QtWidgets.QHBoxLayout(self)

        
        font = QtGui.QFont()
        font.setPointSize(20)

        
        hbox1 = QtWidgets.QVBoxLayout()
        self.label_1 = QtWidgets.QLabel("EDIT EXISTED DATA", self)
        self.label_1.setFont(font)
        hbox1.addWidget(self.label_1)
        
        self.pushButton_1 = QtWidgets.QPushButton("Edit", self)
        self.pushButton_1.setObjectName("edit_button")
        self.pushButton_1.clicked.connect(self.go_to_edit_page)
        self.pushButton_1.setFont(QtGui.QFont("Arial", 20))
        hbox1.addWidget(self.pushButton_1)
        hbox1.setAlignment(QtCore.Qt.AlignCenter)
        hbox1.addSpacing(5) 
        main_layout.addLayout(hbox1)

        
        hbox2 = QtWidgets.QVBoxLayout()
        self.label_2 = QtWidgets.QLabel("CREATE REQUIREMENTS", self)
        self.label_2.setFont(font)
        hbox2.addWidget(self.label_2)
        
        self.pushButton_2 = QtWidgets.QPushButton("Requirement", self)
        self.pushButton_2.setObjectName("requirement_button")
        self.pushButton_2.clicked.connect(self.go_to_requirement_page)
        self.pushButton_2.setFont(QtGui.QFont("Arial", 20))
        hbox2.setAlignment(QtCore.Qt.AlignCenter)
        hbox2.addWidget(self.pushButton_2)
        
        main_layout.addLayout(hbox2)

        
        hbox3 = QtWidgets.QVBoxLayout()
        self.label_3 = QtWidgets.QLabel("PREVIEW EXISTED DATA", self)
        self.label_3.setFont(font)
        hbox3.addWidget(self.label_3)
        
        self.pushButton_3 = QtWidgets.QPushButton("Preview", self)
        self.pushButton_3.setObjectName("preview_button")
        self.pushButton_3.clicked.connect(self.go_to_prev_page)
        self.pushButton_3.setFont(QtGui.QFont("Arial", 20))
        hbox3.addWidget(self.pushButton_3)
        hbox3.setAlignment(QtCore.Qt.AlignCenter)
        main_layout.addLayout(hbox3)

        
        hbox4 = QtWidgets.QVBoxLayout()
        self.label_4 = QtWidgets.QLabel("EDIT EXISTED REQUIREMENT", self)
        self.label_4.setFont(font)
        hbox4.addWidget(self.label_4)
        
        self.pushButton_4 = QtWidgets.QPushButton("Requirement Edit", self)
        self.pushButton_4.setObjectName("requirement_edit_button")
        self.pushButton_4.clicked.connect(self.go_to_reqed_page)
        self.pushButton_4.setFont(QtGui.QFont("Arial", 20))
        hbox4.addWidget(self.pushButton_4)
        hbox4.setAlignment(QtCore.Qt.AlignCenter)
        main_layout.addLayout(hbox4)

       
        self.back_button = QtWidgets.QPushButton('⬅️', self) #go back butonu, emoji eklendi
        self.back_button.clicked.connect(self.go_back)
        self.back_button.setFont(QtGui.QFont("Arial", 16))
        

    def retranslateUi(self): #burasi silinebilir, qtdesigner dan geldi 
        _translate = QtCore.QCoreApplication.translate
        self.pushButton_1.setText(_translate("DecisionPage", "Edit"))
        self.pushButton_2.setText(_translate("DecisionPage", "Requirement"))
        self.pushButton_3.setText(_translate("DecisionPage", "Preview"))
        self.pushButton_4.setText(_translate("DecisionPage", "Requirement Edit"))
        self.label_1.setText(_translate("DecisionPage", "EDIT"))
        self.label_2.setText(_translate("DecisionPage", "REQUIREMENT"))
        self.label_3.setText(_translate("DecisionPage", "PREVIEW"))
        self.label_4.setText(_translate("DecisionPage", "REQUIREMENT EDIT"))

    def go_to_baslangic_page(self): #obur sayfalara gecebilmek icin
        self.stacked_widget.setCurrentIndex(3)
    
    def go_to_prev_page(self):
        self.stacked_widget.setCurrentIndex(15) 
    
    def go_to_edit_page(self):
        self.stacked_widget.setCurrentIndex(14)

    def go_to_requirement_page(self):
        
        self.stacked_widget.setCurrentIndex(12)
    
    def go_to_reqed_page(self):
        self.stacked_widget.setCurrentIndex(13)

    def go_back(self):
        self.stacked_widget.setCurrentIndex(10) #admin signin page ine gidecek


class BaslangicPenceresi(QtWidgets.QWidget):
    def __init__(self, stacked_widget):
        super().__init__()
        self.stacked_widget = stacked_widget
        self.setupUi()

    def setupUi(self):
        layout = QtWidgets.QVBoxLayout()
        
        font = QtGui.QFont()
        font.setPointSize(15)
        
        self.label = QtWidgets.QLabel("Determine Type:", self)
        self.label.setFont(font)  # Label ivin font ayari
        layout.addWidget(self.label)

        self.combobox = QtWidgets.QComboBox(self)
        self.combobox.addItem("Type 1")
        self.combobox.addItem("Type 2")
        self.combobox.addItem("Type 3")
        self.combobox.setFont(font)  # combobox font ayari
        layout.addWidget(self.combobox)

        self.continue_button = QtWidgets.QPushButton("Continue", self)
        self.continue_button.setFont(font)  # buton için font büyüklüğü ayarı
        self.continue_button.clicked.connect(self.continue_to_giris_penceresi)
        layout.addWidget(self.continue_button)
        
        self.back_button = QtWidgets.QPushButton('Go Back', self)
        self.back_button.setFont(font)  
        self.back_button.clicked.connect(self.go_back)
        layout.addWidget(self.back_button)

        self.setLayout(layout)

    def get_combobox_value(self):
        return self.combobox.currentText()

    def continue_to_giris_penceresi(self):
        selected_option = self.combobox.currentText()
        if selected_option in ["Type 1", "Type 2", "Type 3"]:
            self.stacked_widget.setCurrentIndex(4)
            self.stacked_widget.widget(4).set_type(selected_option)
        else:
            QtWidgets.QMessageBox.warning(self, "Warning", "Please select a valid option.")

    def go_back(self):
        self.stacked_widget.setCurrentIndex(2)  

class GirisPencere(QtWidgets.QWidget): 
    
    def __init__(self, stacked_widget):
        super().__init__()
        self.stacked_widget = stacked_widget
        self.type = ""
        self.setupUi()

    def setupUi(self):
        self.setWindowTitle("MainWindow")
        self.resize(558, 308)  

        font = QtGui.QFont() #font ayarlamak icin eklendi
        font.setPointSize(20)

        
        main_layout = QtWidgets.QVBoxLayout(self)

        
        self.label = QtWidgets.QLabel(self.type, self)
        self.label.setFont(font)
        self.label.setAlignment(QtCore.Qt.AlignCenter)
        main_layout.addWidget(self.label, alignment=QtCore.Qt.AlignCenter)

        
        hbox1 = QtWidgets.QVBoxLayout()
        self.label_1 = QtWidgets.QLabel("DETERMINE DETAIL WINDOW", self)
        self.label_1.setFont(font)
        hbox1.addWidget(self.label_1)

        self.pushButton_1 = QtWidgets.QPushButton("Detail Window", self)
        self.pushButton_1.setObjectName("edit_button")
        self.pushButton_1.clicked.connect(self.go_to_detail_window)
        self.pushButton_1.setFont(QtGui.QFont("Arial", 20))
        hbox1.addWidget(self.pushButton_1)
        hbox1.setAlignment(QtCore.Qt.AlignCenter)
        hbox1.addSpacing(5)
        main_layout.addLayout(hbox1)

        
        hbox2 = QtWidgets.QVBoxLayout()
        self.label_2 = QtWidgets.QLabel("DETERMINE LIST WINDOW", self)
        self.label_2.setFont(font)
        hbox2.addWidget(self.label_2)

        self.pushButton_2 = QtWidgets.QPushButton("List Window", self)
        self.pushButton_2.setObjectName("requirement_button")
        self.pushButton_2.clicked.connect(self.go_to_list_window)
        self.pushButton_2.setFont(QtGui.QFont("Arial", 20))
        hbox2.setAlignment(QtCore.Qt.AlignCenter)
        hbox2.addWidget(self.pushButton_2)

        main_layout.addLayout(hbox2)

       
        hbox3 = QtWidgets.QVBoxLayout()
        self.label_3 = QtWidgets.QLabel("DETERMINE SHORT INFO WINDOW", self)
        self.label_3.setFont(font)
        hbox3.addWidget(self.label_3)

        self.pushButton_3 = QtWidgets.QPushButton("Short Info Window", self)
        self.pushButton_3.setObjectName("shortinfo_button")
        self.pushButton_3.clicked.connect(self.go_to_short_info_window)
        self.pushButton_3.setFont(QtGui.QFont("Arial", 20))
        hbox3.addWidget(self.pushButton_3)
        hbox3.setAlignment(QtCore.Qt.AlignCenter)
        main_layout.addLayout(hbox3)


        self.back_button = QtWidgets.QPushButton('⬅️', self)
        self.back_button.clicked.connect(self.go_back)
        self.back_button.setFont(QtGui.QFont("Arial", 16))
        


    def go_to_list_window(self): #list window'a gitme (isimleri degisecek) 
        self.stacked_widget.setCurrentIndex(6)

    def go_to_detail_window(self): 
        num, ok = QInputDialog.getInt(self, "Determine Fields", "Enter number of fields:")
        if ok:
            self.stacked_widget.widget(5).set_num(num)
            self.stacked_widget.setCurrentIndex(5)
            self.stacked_widget.widget(4).set_type(self.type)

    def go_to_short_info_window(self):
        self.stacked_widget.setCurrentIndex(7)

    def set_type(self, type):
        self.type = type
        self.label.setText(self.type)   # Type bilgisini sakladik

    def go_back(self):
        self.stacked_widget.setCurrentIndex(3)

    #def resizeEvent(self, event): #bura silinecek
        #super().resizeEvent(event)
    

class SecondPencere(QtWidgets.QWidget):  
    def __init__(self, stacked_widget):
        super().__init__()
        self.stacked_widget = stacked_widget
        self.num = 0 
        self.setupUi()
        self.location_label = None
        self.baslangic_page = stacked_widget.baslangic_page 

    def setupUi(self): 
        self.layout = QVBoxLayout(self) 

        

        self.scroll_area = QScrollArea(self) #scroll icin 
        self.scroll_area.setWidgetResizable(True) 
        self.scroll_content = QWidget() 
        self.scroll_layout = QVBoxLayout(self.scroll_content)

        self.title_of_all = QLabel("Enter title for page:")
        self.title_of_all.setFont(QtGui.QFont("Arial", 12))
        self.scroll_layout.addWidget(self.title_of_all)
        
        
        self.title_input = [] 
        
        self.page_titles = QLineEdit()
        self.page_titles.setFont(QtGui.QFont("Arial", 12))
        self.scroll_layout.addWidget(self.page_titles) #scroll latout inputların dışındaki kısım için kullanıldı
        self.title_input.append(self.page_titles)

        self.label = QLabel('', self.scroll_content)
        self.label.setFont(QtGui.QFont("Arial", 12))
        self.scroll_layout.addWidget(self.label)


        self.input_layout = QVBoxLayout()
        self.scroll_layout.addLayout(self.input_layout)
        
        
        

        self.scroll_area.setWidget(self.scroll_content)
        self.layout.addWidget(self.scroll_area)

        

        self.enum_label = QLabel("Enter desired number for enum types:", self)
        self.enum_label.setFont(QtGui.QFont("Arial", 12))
        self.input_layout.addWidget(self.enum_label)
        self.enum_label.hide() 

        #self.list_widget = QListWidget()
        #self.layout.addWidget(self.list_widget)

        self.name_inputs = [] #girilenleri tutuyoruz
        self.type_comboboxes = []
        self.min_inputs = []
        self.max_inputs = []
        self.default_inputs = []
        self.name_labels = []
        self.type_labels = []
        self.checkboxes = []
        self.loc_checkboxes = []
        self.string_start_checkboxes = []
        self.string_char_checkboxes = []
        self.string_punc_checkboxes = []
        #self.title_input = []
        #self.enum_line_input = []
        self.line_edits = []



        

        self.submit_button = QPushButton("Save", self)
        self.submit_button.setFont(QtGui.QFont("Arial", 12))
        self.submit_button.clicked.connect(self.get_names)
        self.layout.addWidget(self.submit_button)

        self.req_save_button = QPushButton("Save Requirement", self)
        self.req_save_button.setFont(QtGui.QFont("Arial", 12))
        self.req_save_button.clicked.connect(self.save_requirement)
        self.layout.addWidget(self.req_save_button)

        self.back_button = QPushButton('Go Back', self)
        self.back_button.clicked.connect(self.go_back_to_first_page)
        self.back_button.setFont(QtGui.QFont("Arial", 12))
        
        self.layout.addWidget(self.back_button)
       
        """self.radio_button_edit = QRadioButton("Editable", self)  #secondpencere de bu butona ihtiyaç olmadığı için kullanmadık
        self.radio_button_edit.setChecked(False)  
        self.radio_button_edit.toggled.connect(self.radiobutton_editable_check)
        self.layout.addWidget(self.radio_button_edit)

        self.radio_button_nonedit = QRadioButton("Non-Editable", self)
        
        self.radio_button_nonedit.toggled.connect(self.radiobutton_noneditable_check)
        self.layout.addWidget(self.radio_button_nonedit)

        self.save_checkbox = QCheckBox("Save Button will be added", self)
        self.layout.addWidget(self.save_checkbox)
        self.save_checkbox.hide()

        self.edit_checkbox = QCheckBox("Edit Button will be added", self)
        self.layout.addWidget(self.edit_checkbox)
        self.edit_checkbox.hide()

        self.delete_checkbox = QCheckBox("Delete Button will be added", self)
        self.layout.addWidget(self.delete_checkbox)
        self.delete_checkbox.hide()"""
    
    
        
        
        # excel dosyasi buradan gelecek (hata aldim duzeltmek amacli)
        self.df = pd.read_excel('requirement_templates.xlsx', engine='openpyxl')

    def set_num(self, num): 
        
        self.num = num
       
        self.label.setText(f"Give {self.num} fields to name and types:") #verilen sayi kadar alan olusturuldu
        
        self.create_inputs() 

    def create_inputs(self): 
        
    
        for i in reversed(range(len(self.name_inputs))): #girilenleri silme, arayuzde gozukmez sonrasinda
            self.input_layout.removeWidget(self.name_inputs[i])
            self.name_inputs[i].deleteLater() 
            self.input_layout.removeWidget(self.type_comboboxes[i])
            self.type_comboboxes[i].deleteLater()
            self.input_layout.removeWidget(self.name_labels[i])
            self.name_labels[i].deleteLater()
            self.input_layout.removeWidget(self.type_labels[i])
            self.type_labels[i].deleteLater()
            
            if len(self.min_inputs) > i: #eger min_inputs'in uzunlugu i'den buyukse
                self.input_layout.removeWidget(self.min_inputs[i]) #min degeri silme
                self.min_inputs[i].deleteLater()
                self.input_layout.removeWidget(self.max_inputs[i]) #max degeri silme
                self.max_inputs[i].deleteLater()
                self.input_layout.removeWidget(self.default_inputs[i]) #default degeri silme
                self.default_inputs[i].deleteLater()
                if len(self.loc_checkboxes) > i:
                    self.input_layout.removeWidget(self.loc_checkboxes[i])
                    self.loc_checkboxes[i].deleteLater()

        self.name_inputs = [] 
        self.type_comboboxes = []
        self.min_inputs = []
        self.max_inputs = []
        self.default_inputs = []
        self.name_labels = []
        self.type_labels = []
        self.checkboxes = []
        self.loc_checkboxes = []
        self.enum_inputs = []
        self.string_start_checkboxes = []
        self.string_char_checkboxes = []
        self.string_punc_checkboxes = []
        self.str_spinbox_input = []
        self.str_spinbox_input = []
        #self.title_input = []
        #self.enum_line_input = []
        self.req_button_info = []
        #kaydedebilmek icin buralarda depolandi
        

        for i in range(self.num): #verilen sayi kadar alan olusturuldu
            name_label = QLabel(f"Name of the field {i+1}:", self)
            self.input_layout.addWidget(name_label)    
            self.name_labels.append(name_label)
            name_label.setFont(QtGui.QFont("Arial", 12))

            name_input = QLineEdit(self) 
            self.input_layout.addWidget(name_input)
            self.name_inputs.append(name_input)
            name_input.setFont(QtGui.QFont("Arial", 12))

            type_label = QLabel(f"Type of the field {i+1}:", self)
            self.input_layout.addWidget(type_label)    
            self.type_labels.append(type_label)
            type_label.setFont(QtGui.QFont("Arial", 12))

            type_combobox = QComboBox(self)
            type_combobox.addItem("")
            type_combobox.addItem("string")
            type_combobox.addItem("integer")
            type_combobox.addItem("double")
            type_combobox.addItem("location")
            type_combobox.addItem("enum")
            type_combobox.setCurrentIndex(0)
            type_combobox.setFont(QtGui.QFont("Arial", 12))
            
            type_combobox.currentIndexChanged.connect(lambda index, i=i: self.on_type_change(index, i))
            self.input_layout.addWidget(type_combobox)
            self.type_comboboxes.append(type_combobox)

            #self.enum_line_input.append([])

            loc_checkbox = QCheckBox("Multi location.", self)
            loc_checkbox.stateChanged.connect(lambda state_loc, i=i: self.loc_checkbox_change(state_loc, i))
            loc_checkbox.hide()
            loc_checkbox.setFont(QtGui.QFont("Arial", 12))
            self.input_layout.addWidget(loc_checkbox)
            self.loc_checkboxes.append(loc_checkbox)

            start_num_checkbox =  QCheckBox("Start with number", self)
            start_num_checkbox.hide()
            start_num_checkbox.setFont(QtGui.QFont("Arial", 12))
            self.input_layout.addWidget(start_num_checkbox)
            self.string_start_checkboxes.append(start_num_checkbox)


            string_character_checkbox = QCheckBox("Turkish character allowed", self)
            string_character_checkbox.hide()
            string_character_checkbox.setFont(QtGui.QFont("Arial", 12))
            self.input_layout.addWidget(string_character_checkbox)
            self.string_char_checkboxes.append(string_character_checkbox)

            string_punctuation_checkbox = QCheckBox("Punctuation allowed", self)
            string_punctuation_checkbox.hide()
            string_punctuation_checkbox.setFont(QtGui.QFont("Arial", 12))
            self.input_layout.addWidget(string_punctuation_checkbox)
            self.string_punc_checkboxes.append(string_punctuation_checkbox)

            self.requirement_button = QPushButton("Choose Requirement", self)
            self.requirement_button.hide()
            self.requirement_button.setFont(QtGui.QFont("Arial", 12))
            self.input_layout.addWidget(self.requirement_button)
            self.req_button_info.append(self.requirement_button)
            self.requirement_button.clicked.connect(self.choose_requirement)
            

            enum_spinbox = QSpinBox()
            enum_spinbox.setMinimum(0)
            enum_spinbox.setMaximum(100000)
            enum_spinbox.setSingleStep(1.0)
            enum_spinbox.setSpecialValueText("Choose enum type number")
            enum_spinbox.hide() #baslangicta gozukmemesi icin
            enum_spinbox.setFont(QtGui.QFont("Arial", 12))
            self.input_layout.addWidget(enum_spinbox)
            self.enum_inputs.append(enum_spinbox) #hata olabilir


            str_spinbox = QSpinBox() 
            str_spinbox.setMinimum(0)
            str_spinbox.setMaximum(100000)
            str_spinbox.setSingleStep(1)
            str_spinbox.setSpecialValueText("Give max character count")
            str_spinbox.setFont(QtGui.QFont("Arial", 12))
            str_spinbox.hide()
            self.input_layout.addWidget(str_spinbox)
            self.str_spinbox_input.append(str_spinbox)
            
        
            min_input = QSpinBox() if type_combobox.currentText() == "integer" else QDoubleSpinBox() #integerda spinbox, double'da double spinbox
            min_input.setMinimum(0)
            min_input.setSingleStep(1.0 if type_combobox.currentText() == "integer" else 0.1)
            min_input.setSpecialValueText("Choose minimum value")
            min_input.setFont(QtGui.QFont("Arial", 12))
            min_input.hide()
            self.input_layout.addWidget(min_input)
            self.min_inputs.append(min_input)

            max_input = QSpinBox() if type_combobox.currentText() == "integer" else QDoubleSpinBox()
            max_input.setMaximum(10000000)
            max_input.setSingleStep(1.0 if type_combobox.currentText() == "integer" else 0.1)
            max_input.setSpecialValueText("Choose maximum value")
            max_input.setFont(QtGui.QFont("Arial", 12))
            max_input.hide()
            self.input_layout.addWidget(max_input)
            self.max_inputs.append(max_input)

            default_input = QDoubleSpinBox()
            default_input.setSpecialValueText("Choose a default value")
            default_input.setFont(QtGui.QFont("Arial", 12))
            default_input.hide()
            self.input_layout.addWidget(default_input)
            self.default_inputs.append(default_input)

            checkbox = QCheckBox("Determine min, max, default value.", self)
            checkbox.stateChanged.connect(lambda state, i=i: self.on_checkbox_change(state, i))
            checkbox.setFont(QtGui.QFont("Arial", 12))
            checkbox.hide()
            self.input_layout.addWidget(checkbox)
            self.checkboxes.append(checkbox)

        
  #buralar enum icindi calismadi:
    #def enum_input_dialog(self):
      #  num, ok = QInputDialog.getInt(self, "Input Dialog", "Enter number of QLineEdits:", min=1, max=10000)
      #  if ok:
        # Find the index of the current name input
      #      current_index = len(self.enum_line_input) - 1
      #      self.create_enum_line(num, current_index)

    #def create_enum_line(self, num, index):
     #   for _ in range(num):
     #       enum_line = QLineEdit(self)
     #       self.enum_line_input[index].append(enum_line)
     #       self.input_layout.insertWidget(self.scroll_layout.count() - 2, enum_line)  # Insert before buttons  

    def choose_requirement(self):
        dummy_line_edit = QLineEdit()
        self.line_edit_width = dummy_line_edit.sizeHint().width()
        self.line_edit_height = dummy_line_edit.sizeHint().height()

        try:
            
            df = pd.read_excel('requirement_templates.xlsx', sheet_name='Sheet1', engine='openpyxl') #requirement_template den aldik 
            requirement_templates = df['Template Name'].unique().tolist() #unique olanlari aldik

            
            dialog = RequirementDialog(requirement_templates, self) #requirement dialog olusturuldu
            if dialog.exec_() == QDialog.Accepted: #dialog acildi mi
                
                
                selected_requirement = dialog.get_selected_requirement() #secilen requirement alindi
                self.filtered_data = df[df['Template Name'] == selected_requirement] #secilen requirement'a gore filtreleme yapildi, secilen requirement in template name i acildi

                if not self.filtered_data.empty:
                    #opsiyonel olmayanlar otomatik gelmesi
                    non_optional_texts = self.filtered_data[self.filtered_data['Optional'] == False]

                    for _, row in non_optional_texts.iterrows():
                        self.add_text_segments(row)

                    # opsiyonel olanlari user karar verecek gelip gelmeyeceklerini, opsiyonel olmayanlar direk gelecek
                    texts = self.filtered_data['Text'].tolist()
                    dialog = TextSelectionDialog(texts, self)
                    if dialog.exec_() == QDialog.Accepted: 
                        selected_texts = dialog.get_selected_texts() #secilen textler alindi
                        if selected_texts:
                            for _, row in self.filtered_data.iterrows(): 
                                if row['Text'] in selected_texts:  #secilen textlerin icinde olanlar eklendi
                                    self.add_text_segments(row)
                        else:
                            QMessageBox.warning(self, "Selection", "No text selected.") #error handling
                    else:
                        QMessageBox.warning(self, "Selection", "No text selected.")
                else:
                    QMessageBox.warning(self, "Selected Template", "No texts found for the selected template.")
            else:
                QMessageBox.information(self, "Selection", "No requirement template selected.")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error occurred: {str(e)}")

    def add_text_segments(self, row): #text segmentlerini eklemek icin
        sentence_label = QLabel(row['Sentence']) #sentence label olusturuldu
        sentence_label.setFont(QtGui.QFont("Arial", 13)) #font ayari

        #sentence_name = QLabel(row['Template Name']) #template name label olusturuldu

        template_name_label = QLabel(row['Template Name'])
        template_name_label.setFont(QtGui.QFont("Arial", 13))

        description = QLabel("Give values to parameters:")
        description.setFont(QtGui.QFont("Arial", 13))

        text = str(row['Text']) #text alindi
        parameters = [param.strip() for param in str(row["Parameters"]).split(",")] #parametreler alindi

        text_segments = [] #text segmentleri olusturuldu
        remaining_text = text #kalan text

        for param in parameters:
            param_index = remaining_text.find(param) #parametre indexi
            if param_index != -1: #eger parametre varsa
                if param_index > 0: #eger index 0 dan buyukse
                    text_segments.append(remaining_text[:param_index]) #text segmente ekle
                text_segments.append(param) #parametre ekle
                remaining_text = remaining_text[param_index + len(param):] #kalan texti al

        if remaining_text: #eger kalan text varsa
            text_segments.append(remaining_text) #text segmente ekle

        row_edit = [] 
        self.parameter_widgets = {}

        text_layout = QHBoxLayout()
        text_layout.setContentsMargins(0, 0, 0, 0)
        text_layout.setSpacing(0)
        text_layout.setAlignment(Qt.AlignLeft)

        
        for segment in text_segments:
            if segment.strip() in parameters: #eger segment parametrelerde varsa
                line_edit = QLineEdit(self)
                line_edit.setPlaceholderText(segment)
                line_edit.setFixedSize(self.line_edit_width, self.line_edit_height)
                font = line_edit.font()
                font.setPointSize(13)
                font.setItalic(True)
                font.setUnderline(True)
                line_edit.setFont(font)
                line_edit.setStyleSheet("background-color: lightblue;")
                text_layout.addWidget(line_edit)
                row_edit.append(line_edit)
                line_edit.setReadOnly(True)
            else:
                words = segment.split() #segmenti kelimelere ayir, parameter lerden kalan
                for word in words:
                    if word.strip():
                        line_edit = QLineEdit(self)
                        line_edit.setText(word)
                        line_edit.setFixedSize(self.line_edit_width, self.line_edit_height)
                        font = line_edit.font()
                        font.setPointSize(13)
                        line_edit.setFont(font)
                        text_layout.addWidget(line_edit)
                        row_edit.append(line_edit)

        param_layout = QHBoxLayout() #parametre layoutu olusturuldu
        param_layout.setContentsMargins(0, 0, 0, 0)
        param_layout.setSpacing(0)
        param_layout.setAlignment(Qt.AlignLeft)

        self.param_value_list = []

        for param in parameters: #parametreler icin
            param_label = QLabel(param + ":", self) #parametre label olusturuldu
            param_label.setFont(QtGui.QFont("Arial", 13))
            param_label.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
            
            param_edit = QLineEdit(self) 
            param_edit.setFixedSize(self.line_edit_width, self.line_edit_height)
            param_edit.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
            param_edit.setStyleSheet("background-color: lightyellow")
            font = param_edit.font()
            font.setPointSize(13)
            param_edit.setFont(font)


            self.param_value_list.append((param, param_edit))
            
            param_edit.textChanged.connect(lambda text, p=param: self.update_parameter_text(p, text)) #parametre texti guncelleme
            
            param_layout.addWidget(param_label)
            param_layout.addWidget(param_edit)
            self.parameter_widgets[param] = param_edit

        text_widget = QWidget()
        text_widget.setLayout(text_layout)

        parameter_widget = QWidget()
        parameter_widget.setLayout(param_layout)

        sender_button = self.sender() #sender button
        position = self.input_layout.indexOf(sender_button) + 1 #position alindi
        self.input_layout.insertWidget(position-2, template_name_label) #template name eklendi
        self.input_layout.insertWidget(position-1, description) #description eklendi
        self.input_layout.insertWidget(position, sentence_label) #sentence label eklendi
        self.input_layout.insertWidget(position + 1, text_widget) #text widget eklendi
        self.input_layout.insertWidget(position + 2, parameter_widget) #parametre widget eklendi
        self.line_edits.append((position, row_edit))

        
    def update_parameter_text(self, parameter, text): #parametre texti guncelleme
        
        for pos, row_edit in self.line_edits: #row edit icindeki line editler icin
            for line_edit in row_edit: #row edit icindeki line editler icin
                if line_edit.placeholderText() == parameter: #eger line editin placeholder texti parametreyle ayniysa
                    line_edit.setText(text) #texti guncelle

    def get_param_value(self):
        for param, param_edit in self.param_value_list: #parametre ve parametre edit icin
            param_value = param_edit.text() #parametre editin textini al
            for pos, row_edit in self.line_edits: #row edit icin
                for line_edit in row_edit: #line edit icin
                    if line_edit.placeholderText() == param: #eger line editin placeholder texti parametreyle ayniysa
                        line_edit.setText(param_value) #texti guncelle
    

    def save_requirement(self):
        try:
            file_path = 'requirements.xlsx'
            try:
                existing_df = pd.read_excel(file_path, engine='openpyxl')
            except FileNotFoundError:
                existing_df = pd.DataFrame()  # yoksa dosya olusturuldı (bu daha erken bi yere eklenmeli hata verebilir)

            new_data = []
            self.param_values = {}  #parametreler store

            for param, param_edit in self.param_value_list:
                param_value = param_edit.text()  # parametrelerin text degeri alindi, icine yazilanlar
                self.param_values[param] = param_value  #parametreler dictionary ye kaydedildi 

            
            if hasattr(self, 'filtered_data'):
                template_names = self.filtered_data['Template Name'].unique().tolist()  # template name ler geldi filtered data verisi cekilerek
            else:
                template_names = [] #bos kalmali eger yoksa

            
            type_info = self.baslangic_page.get_combobox_value() #type bilgisini getirdik

            # gecici olarak data_info dictionary olusturuldu
            self.data_info = {}

            for i in range(self.num): #verilen sayi kadar
                name = self.name_inputs[i].text() #name alindi
                data_type = self.type_comboboxes[i].currentText() #data type alindi
                title = self.page_titles.text() #title alindi

                if title not in self.data_info:
                    self.data_info[title] = {} #title yoksa olusturuluyor ama bu silinebilir cunku title hep olusturuluyor

                if data_type in ["integer", "double"]: #burasi onceki koddada var, girilen verileri kaydetmek icin
                    if self.checkboxes[i].isChecked():
                        min_value = self.min_inputs[i].value()
                        max_value = self.max_inputs[i].value()
                        default_value = self.default_inputs[i].value()
                        checkbox_value = self.checkboxes[i].isChecked()

                        self.data_info[title][name] = {
                            "data type": data_type,
                            "min": min_value,
                            "max": max_value,
                            "default": default_value,
                            "option_enabled": checkbox_value,
                            "Type": type_info
                        }
                    else:
                        self.data_info[title][name] = {
                            "data type": data_type,
                            "min": "Not Determined",
                            "max": "Not Determined",
                            "default": "Not Determined",
                            "option_enabled": False,
                            "Type": type_info
                        }
                elif data_type in ["location"]:
                    if self.loc_checkboxes[i].isChecked():
                        loc_checkbox_value = self.loc_checkboxes[i].isChecked()
                        loc_info = self.location_label.text() if self.location_label else None

                        self.data_info[title][name] = {
                            "data type": data_type,
                            "loc_option_enabled": loc_checkbox_value,
                            "loc_info_enabled": loc_info,
                            "Type": type_info
                        }
                    else:
                        self.data_info[title][name] = {
                            "data type": data_type,
                            "loc_option_enabled": False,
                            "loc_info_enabled": "Not Determined",
                            "Type": type_info
                        }
                elif data_type in ["string"]:
                    char_num = self.str_spinbox_input[i].value()

                    if self.string_start_checkboxes[i].isChecked() or self.string_char_checkboxes[i].isChecked() or self.string_punc_checkboxes[i].isChecked(): #string checkbox lari kontrol ediyor (bunlarin konumunda sorun var)
                        start_num = self.string_start_checkboxes[i].isChecked()
                        character_choice = self.string_char_checkboxes[i].isChecked()
                        punctuation_choice = self.string_punc_checkboxes[i].isChecked()

                        self.data_info[title][name] = {
                            "data type": data_type,
                            "start_num": start_num,
                            "character_choice": character_choice,
                            "punctuation_choice": punctuation_choice,
                            "character_number": char_num,
                            "Type": type_info
                        }
                    else:
                        self.data_info[title][name] = {
                            "data type": data_type,
                            "start_num": False,
                            "character_choice": False,
                            "punctuation_choice": False,
                            "Type": type_info
                        }
                elif data_type in ["enum"]:
                    number = self.enum_inputs[i].value()
                    self.data_info[title][name] = {
                        "data type": data_type,
                        "enum type number": number,
                        "Type": type_info
                    }

            
            for pos, row_edit in self.line_edits:
                for line_edit in row_edit:
                    placeholder = line_edit.placeholderText()
                    text = line_edit.text()
                    if placeholder and text:
                        new_data.append({
                            'Category': f"{title} - {name} - Text Segment", #bu sekilde gozukecek excel e kaydedildiginde.
                            'Value': f"{placeholder}: {text}" 
                        })

                self.phrases = [line_edit.text() for line_edit in row_edit if line_edit.text()]
                if self.phrases:
                    new_data.append({
                        'Category': f"{title} - {name} - Sentence", #sentence kaydedilmesi için
                        'Value': ' '.join(self.phrases)
                    })

            
            for param, value in self.param_values.items():
                new_data.append({
                    'Category': f"{title} - {name} - Parameter", 
                    'Value': f"{param}: {value}"
                })

            
            for template in template_names:
                new_data.append({
                    'Category': f"{title} - {name} - Template Name",
                    'Value': template
                })

            
            for title, fields in self.data_info.items():
                for name, attributes in fields.items():
                    for key, value in attributes.items():
                        new_data.append({
                            'Category': f"{title} - {name} - {key}", #diger bilgiler key ile gelir bu sekilde kaydolur 
                            'Value': value
                        })

            df_new = pd.DataFrame(new_data)

            if not existing_df.empty:
                df_combined = pd.concat([existing_df, df_new], axis=1) #eskiler silinmemesi icin 
            else:
                df_combined = df_new #eger yeni data yoksa eskiler kalmasi icin

            df_combined.to_excel(file_path, index=False, engine='openpyxl') 

            QMessageBox.information(self, "Success", "Data saved to Excel successfully.")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error occurred while saving to Excel: {str(e)}")


    def get_names(self):
        
        self.names = {}

        type_info = self.baslangic_page.get_combobox_value() #type bilgisini alabilmek icin 
        
       
        for i in range(self.num):
            name = self.name_inputs[i].text()
            data_type = self.type_comboboxes[i].currentText()
            title = self.page_titles.text()

            if data_type in ["integer", "double"]:
                if self.checkboxes[i].isChecked(): 
                   
                    min_value = self.min_inputs[i].value()
                    max_value = self.max_inputs[i].value()
                    default_value = self.default_inputs[i].value()
                    checkbox_value = self.checkboxes[i].isChecked()
                    
                    self.names[name] = {
                        "page title": title,
                        "data type": data_type,
                        "min": min_value,
                        "max": max_value,
                        "default": default_value,
                        "option_enabled": checkbox_value,
                        "Type": type_info
                        
                    }  
                    
                else:  # Determine checkbox secilmedigi zaman
                   
                    self.names[name] = {
                        "page title": title,
                        "data type": data_type,
                        "min": "Not Determined",
                        "max": "Not Determined",
                        "default": "Not Determined",
                        "option_enabled": False,
                        "Type": type_info
                        
                    }
            elif data_type in ["location"]:
                    if self.loc_checkboxes[i].isChecked():
                        loc_checkbox_value = self.loc_checkboxes[i].isChecked()
                        loc_info = self.location_label.text() if self.location_label else None
                        
                        self.names[name] = {
                            "page title": title,
                            "data type": data_type,
                            "loc_option_enabled": loc_checkbox_value,
                            "loc_info_enabled": loc_info,
                            "Type": type_info
                        }
                    else:
                        self.names[name] = {
                            "page title": title,
                            "data type": data_type,
                            "loc_option_enabled": False,
                            "loc_info_enabled": "Not Determined",
                            "Type": type_info
                        }
            
            elif data_type in ["string"]: #in yerine == kullanilabilir 
                char_num = self.str_spinbox_input[i].value()
               
                if self.string_start_checkboxes[i].isChecked() or self.string_char_checkboxes[i].isChecked() or self.string_punc_checkboxes[i].isChecked: #string icin
                    
                    start_num = self.string_start_checkboxes[i].isChecked()
                    character_choice = self.string_char_checkboxes[i].isChecked()
                    punctuation_choice = self.string_punc_checkboxes[i].isChecked()
                    self.names[name] = {"data type": data_type}
                    self.names[name] = {
                        "page title": title,
                        "data type": data_type,
                        "start_num" : start_num,
                        "character_choice": character_choice,
                        "punctuation_choice": punctuation_choice,
                        "character_number": char_num,
                        "Type": type_info 
                    }
                else:
                    self.names[name] = {"data type": data_type}
                    self.names[name] = {
                        "page title": title,
                        "data type": data_type,
                        "start_num" : False,
                        "character_choice": False,
                        "punctuation_choice": False,
                        "Type": type_info 
                    }

                
            elif data_type in ["enum"]:
                
                number = self.enum_inputs[i].value()
                self.names[name] = {"data type": data_type}
                self.names[name] = {
                    "page title": title,
                    "data type": data_type,
                    "enum type number": number,
                    "Type": type_info 
                }

        #sonrasinda vazgecildi bu checkboxlardan:
       # self.names["Customization"] = {
        #"Editable": self.radio_button_edit.isChecked(),
       # "Non-Editable": self.radio_button_nonedit.isChecked()
       # }

       # self.names["Checkbox Preference"] = {
        #    "Save Button": self.save_checkbox.isChecked(),
       #     "Edit Button": self.edit_checkbox.isChecked(),
       #     "Delete Button": self.delete_checkbox.isChecked()
       # }

        
        
        #self.names[name]["Type"] = type_info
        df = pd.DataFrame.from_dict(self.names, orient='index')

        file_path, _ = QtWidgets.QFileDialog.getSaveFileName(self, 'Save File', "", "Excel Files (*.xlsx)")
        if file_path:
            df.to_excel(file_path, index_label='Field Names')

        QMessageBox.information(self, "Info", "Data saved to Excel.")
        self.back_button.setEnabled(True)

    def on_type_change(self, index, i): 
        
        if self.type_comboboxes[i].currentText() in ["integer", "double"]: #integer ya da double secildiginde min max default gozukebilmesi icin
            self.req_button_info[i].show()
            if self.checkboxes[i].isChecked():
                self.min_inputs[i].show()
                self.max_inputs[i].show()
                self.default_inputs[i].show()
            else:
                self.min_inputs[i].hide()
                self.max_inputs[i].hide()
                self.default_inputs[i].hide()
            #self.min_inputs[i].show()
            #self.max_inputs[i].show()
            self.checkboxes[i].show()
            #self.default_inputs[i].show()
            self.min_inputs[i].setValue(0)
            self.max_inputs[i].setValue(0) 
            self.default_inputs[i].setValue(0)
        else:
            self.min_inputs[i].hide()
            self.max_inputs[i].hide()
            self.checkboxes[i].hide()
            self.default_inputs[i].hide()

        if self.type_comboboxes[i].currentText() == "location":
            self.req_button_info[i].show()
            self.loc_checkboxes[i].show()
        else:
            self.loc_checkboxes[i].hide()

        if self.type_comboboxes[i].currentText() == "string":
            self.req_button_info[i].show()
            self.string_start_checkboxes[i].show()
            self.string_char_checkboxes[i].show()
            self.string_punc_checkboxes[i].show()
            self.str_spinbox_input[i].show()
        else:
            self.string_start_checkboxes[i].hide()
            self.string_char_checkboxes[i].hide()
            self.string_punc_checkboxes[i].hide()
            self.str_spinbox_input[i].hide()

        

        if self.type_comboboxes[i].currentText() in ["integer"]:
            self.min_inputs[i].setSingleStep(1.0) #integer icin 1 artacak
            self.max_inputs[i].setSingleStep(1.0)
        elif self.type_comboboxes[i].currentText() in ["double"]:
            self.min_inputs[i].setSingleStep(0.1) #double icin 0.1 artacak
            self.max_inputs[i].setSingleStep(0.1)
        else:
            self.min_inputs[i].hide()
            self.max_inputs[i].hide()
            self.default_inputs[i].hide()
        
        if self.type_comboboxes[i].currentText() == "enum":
            self.req_button_info[i].show()
            #self.enum_input_dialog()
            self.enum_inputs[i].show()
        else:
            self.enum_inputs[i].hide()  

    def on_checkbox_change(self, state, i):
        if state == Qt.Checked:
            self.min_inputs[i].show()
            self.max_inputs[i].show()
            self.default_inputs[i].show()
        else:
            self.min_inputs[i].hide()
            self.max_inputs[i].hide()
            self.default_inputs[i].hide()

    
    
    def loc_checkbox_change(self, state_loc, i):
        if state_loc == Qt.Checked:
            if self.location_label is None:
                self.location_label = QLabel("Multi location needed.", self)
                #self.input_layout.addWidget(self.location_label)
        else:
            if self.location_label:
                self.location_label.hide()
                self.location_label = None

    
    def radiobutton_editable_check(self, checked):
        if checked:
            self.save_checkbox.show()
            self.edit_checkbox.show()
            self.delete_checkbox.show()
        else:
            self.save_checkbox.hide()
            self.edit_checkbox.hide()
            self.delete_checkbox.hide()

    def radiobutton_noneditable_check(self, checked):
        if checked:
            self.save_checkbox.hide()
            self.edit_checkbox.hide()
            self.delete_checkbox.hide()
        else:
            self.save_checkbox.show()
            self.edit_checkbox.show()
            self.delete_checkbox.show()

    

    def go_back_to_first_page(self): 
       
            for title_input in self.title_input:
                title_input.clear() #title_input olmadığı için hata verebilir
            for name_input in self.name_inputs:
                name_input.clear()
            for type_combobox in self.type_comboboxes:
                type_combobox.setCurrentIndex(4)
            for min_input in self.min_inputs:
                min_input.setValue(4)
            for max_input in self.max_inputs:
                max_input.setValue(4)
            for default_input in self.default_inputs:
                default_input.setValue(4)
            for checkbox in self.checkboxes:
                checkbox.setChecked(False)
            for loc_checkbox in self.loc_checkboxes:
                loc_checkbox.setChecked(False)
            for enum_input in self.enum_inputs:
                enum_input.setValue(4)

            for self.requirement_button in self.req_button_info:
                self.requirement_button.deleteLater()

            #for self.data in self.data_info:
              #  self.data.deleteLater()
           #parameter phrase girişler 

            
           # for self. parameter in self.parameter_widgets:
           #     self.parameter_widgets.clear()
            

           
            if self.location_label:
                self.location_label.hide()
                self.location_label = None

        #requirement ile ilgili olan silme kismi henuz eklenecek

            self.stacked_widget.setCurrentIndex(4)
    #
    # def resizeEvent(self, event):
     #   super().resizeEvent(event)

      #  widget_width = self.width()
       # widget_height = self.height()

       
       # layout_size = self.input_layout.sizeHint()

   
      #  margin_left = (widget_width - layout_size.width()) // 2
      #  margin_top = (widget_height - layout_size.height()) // 2

    # layout'unun pozisyonunu ve boyutu
      #  self.input_layout.setGeometry(QtCore.QRect(margin_left, margin_top, layout_size.width(), layout_size.height()))


class RequirementDialog(QDialog):
    def __init__(self, requirement_templates, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Select Requirement")
        self.layout = QVBoxLayout(self)

        self.list_widget = QListWidget() #liste halinde requirement template isimleri acilir 
        for requirement_template in requirement_templates:
            item = QListWidgetItem(requirement_template)
            self.list_widget.addItem(item)
        self.layout.addWidget(self.list_widget)

        self.button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        self.button_box.accepted.connect(self.accept)
        self.button_box.rejected.connect(self.reject)
        self.layout.addWidget(self.button_box)

    def get_selected_requirement(self):
        selected_item = self.list_widget.currentItem()
        if selected_item:
            return selected_item.text()
        return None

class TextSelectionDialog(QDialog):
    def __init__(self, texts, parent=None):
        super().__init__(parent)
        self.texts = texts
        self.selected_texts = []
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("Select Text")
        self.layout = QVBoxLayout(self)

        self.list_widget = QListWidget()
        self.list_widget.setSelectionMode(QAbstractItemView.MultiSelection) #birden fazla secim yapilabilmesi icin

        df = pd.read_excel('requirement_templates.xlsx', engine='openpyxl')
        text_to_optional = dict(zip(df['Text'], df['Optional'])) # text ve optional dictionary olusturuldu
        
        for text in self.texts:
            optional_value = text_to_optional.get(text, False)
            if optional_value:
                item = QListWidgetItem(text) 
                self.list_widget.addItem(item)
        self.layout.addWidget(self.list_widget)

        self.button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel) #ok ve cancel butonlari 
        self.button_box.accepted.connect(self.accept)
        self.button_box.rejected.connect(self.reject)
        self.layout.addWidget(self.button_box)

    def get_selected_texts(self): #secilen textleri almak icin
        selected_items = self.list_widget.selectedItems() #secilen itemlar
        return [item.text() for item in selected_items] #secilen itemlarin textleri 8kullanici secti)

class ThirdPencere(QtWidgets.QWidget):
    
    def __init__(self, stacked_widget):
        super().__init__()
        self.stacked_widget = stacked_widget
        self.setupUi()

    def setupUi(self): #eklenecek olan gui elemanlari
        self.layout = QVBoxLayout()

        self.third_page_enter = QLabel("Enter title for page", self)
        self.third_page_enter.setFont(QtGui.QFont("Arial", 14)) 
        self.layout.addWidget(self.third_page_enter)

        self.third_page_title = QLineEdit()
        self.third_page_title.setFont(QtGui.QFont("Arial", 12))  
        self.layout.addWidget(self.third_page_title)
        
        self.load_button = QtWidgets.QPushButton('Load Excel File')
        self.load_button.setFont(QtGui.QFont("Arial", 12))  
        self.load_button.clicked.connect(self.load_excel_and_display)
        self.layout.addWidget(self.load_button)
        
        self.table = QtWidgets.QTableWidget()
        self.table.setFont(QtGui.QFont("Arial", 12))  
        self.layout.addWidget(self.table)

        self.edit_checkbox = QtWidgets.QCheckBox("Editable", self)
        self.edit_checkbox.setFont(QtGui.QFont("Arial", 12))  
        self.layout.addWidget(self.edit_checkbox)
        
        self.delete_checkbox = QtWidgets.QCheckBox("Deletable", self)
        self.delete_checkbox.setFont(QtGui.QFont("Arial", 12))  
        self.layout.addWidget(self.delete_checkbox)

        self.save_button = QtWidgets.QPushButton('Save To Excel', self)
        self.save_button.setFont(QtGui.QFont("Arial", 12))  
        self.save_button.clicked.connect(self.save_checkbox_data)
        self.layout.addWidget(self.save_button) 

        self.back_button = QPushButton('Go Back', self)
        self.back_button.setFont(QtGui.QFont("Arial", 12)) 
        self.back_button.clicked.connect(self.go_back_to_first_page)
        self.layout.addWidget(self.back_button)
        
        self.setLayout(self.layout)

    def load_excel_and_display(self):
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(self, 'Open Excel File', '', 'Excel Files (*.xlsx *.xls)')
        if file_path:
            try:
                df = pd.read_excel(file_path, engine='openpyxl')
                self.display_data(df)
            except Exception as e:
                QtWidgets.QMessageBox.critical(self, 'Error', f'Error loading Excel file: {str(e)}')

    def display_data(self, df):
        self.table.setRowCount(df.shape[0])
        self.table.setColumnCount(df.shape[1] + 1)  
        header_labels = list(df.columns) + ['Select'] 

        self.table.setHorizontalHeaderLabels(header_labels)

        for i in range(df.shape[0]):
            for j in range(df.shape[1]):
                item = QtWidgets.QTableWidgetItem(str(df.iloc[i, j]))
                self.table.setItem(i, j, item)

            
            checkbox_item = QtWidgets.QTableWidgetItem()
            checkbox_item.setFlags(QtCore.Qt.ItemIsUserCheckable | QtCore.Qt.ItemIsEnabled)
            checkbox_item.setCheckState(QtCore.Qt.Unchecked)
            self.table.setItem(i, df.shape[1], checkbox_item)

        self.table.setColumnWidth(df.shape[1], 70)

    def save_checkbox_data(self):
        num_rows = self.table.rowCount()
        num_cols = self.table.columnCount() - 1 

        checked_data = []
        for i in range(num_rows):
            if self.table.item(i, num_cols).checkState() == QtCore.Qt.Checked:
                row_data = []
                for j in range(num_cols):
                    item = self.table.item(i, j)
                    if item is not None:
                        row_data.append(item.text())
                    else:
                        row_data.append("")
                checked_data.append(row_data)

        if not checked_data:
            QtWidgets.QMessageBox.warning(self, "Warning", "No rows selected.")
            return

        df = pd.DataFrame(checked_data, columns=list(self.table.horizontalHeaderItem(j).text() for j in range(num_cols)))

        df['Editable'] = 'Add new button will be added' if self.edit_checkbox.isChecked() else ''
        df['Deletable'] = 'Delete button will be added' if self.delete_checkbox.isChecked() else ''
        df['Select'] = 'Selected'  #secilenler gozukecegi icin selected
        df['Title Of Page'] = self.third_page_title.text()

        file_path, _ = QtWidgets.QFileDialog.getSaveFileName(self, 'Save File', "", "Excel Files (*.xlsx)")
        if file_path:
            try:
                df.to_excel(file_path, index=False)
                QtWidgets.QMessageBox.information(self, "Info", "Saved to Excel.")
            except Exception as e:
                QtWidgets.QMessageBox.critical(self, 'Error', f'Error saving Excel file: {str(e)}')

    def go_back_to_first_page(self):
        self.stacked_widget.setCurrentIndex(4)  
        self.table.clearContents()
        self.table.setRowCount(0)
        self.table.setColumnCount(0)

        self.third_page_title.clear()

        self.edit_checkbox.setChecked(False)
        self.delete_checkbox.setChecked(False)

class ForthPencere(QtWidgets.QWidget):
    def __init__(self, stacked_widget):
        super().__init__()
        self.stacked_widget = stacked_widget
        self.setupUi()

    def setupUi(self):
        self.layout = QVBoxLayout()

        self.title_enterance = QLabel("Enter title for page", self)
        self.title_enterance.setFont(QtGui.QFont("Arial", 12))  
        self.layout.addWidget(self.title_enterance)

        self.forth_page_title = QLineEdit()
        self.forth_page_title.setFont(QtGui.QFont("Arial", 12))  
        self.layout.addWidget(self.forth_page_title)
        
        self.load_button = QtWidgets.QPushButton('Open Excel File')
        self.load_button.setFont(QtGui.QFont("Arial", 12))  
        self.load_button.clicked.connect(self.load_excel_and_display)
        self.layout.addWidget(self.load_button)
        
        self.table_widget = QtWidgets.QTableWidget()
        self.table_widget.setFont(QtGui.QFont("Arial", 12))  
        self.layout.addWidget(self.table_widget)

        self.edit_checkbox = QtWidgets.QCheckBox("Editable", self)
        self.edit_checkbox.setFont(QtGui.QFont("Arial", 12))  
        self.layout.addWidget(self.edit_checkbox)
        
        self.location_checkbox = QtWidgets.QCheckBox("Location Button", self)
        self.location_checkbox.setFont(QtGui.QFont("Arial", 12))  
        self.layout.addWidget(self.location_checkbox)

        self.delete_checkbox = QtWidgets.QCheckBox("Delete Button", self)
        self.delete_checkbox.setFont(QtGui.QFont("Arial", 12))  
        self.layout.addWidget(self.delete_checkbox)

        self.save_button = QtWidgets.QPushButton('Save to Excel', self)
        self.save_button.setFont(QtGui.QFont("Arial", 12))  
        self.save_button.clicked.connect(self.save_to_excel)
        self.layout.addWidget(self.save_button) 

        self.back_button = QPushButton('Go Back', self)
        self.back_button.setFont(QtGui.QFont("Arial", 12))  
        self.back_button.clicked.connect(self.go_back_to_first_page)
        self.layout.addWidget(self.back_button)
        
        self.setLayout(self.layout)

    def load_excel_and_display(self):
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(self, 'Open Excel File', '', 'Excel Files (*.xlsx *.xls)')
        if file_path:
            try:
                df = pd.read_excel(file_path, engine='openpyxl')
                self.display_field_names(df)
            except Exception as e:
                QtWidgets.QMessageBox.critical(self, 'Error', f'Error loading Excel file: {str(e)}')

    def display_field_names(self, df):
        field_names = df.columns.tolist()
        self.display_in_table(field_names)

    def display_in_table(self, data):
        self.table_widget.clear()

        num_rows = len(data)
        num_cols = 2  

        self.table_widget.setRowCount(num_rows)
        self.table_widget.setColumnCount(num_cols)

        for i, field_name in enumerate(data):
            item_field_name = QtWidgets.QTableWidgetItem(field_name)
            item_field_name.setFont(QtGui.QFont("Arial", 12))  
            self.table_widget.setItem(i, 0, item_field_name)

            checkbox_item = QtWidgets.QTableWidgetItem()
            checkbox_item.setFlags(QtCore.Qt.ItemIsUserCheckable | QtCore.Qt.ItemIsEnabled)
            checkbox_item.setCheckState(QtCore.Qt.Unchecked)
            self.table_widget.setItem(i, 1, checkbox_item)

        self.table_widget.setHorizontalHeaderLabels(['Field Names', 'Select'])

    def save_to_excel(self):
        file_path, _ = QtWidgets.QFileDialog.getSaveFileName(self, 'Save Data to Excel', '', 'Excel Files (*.xlsx *.xls)')
        if file_path:
            try:
                data = {
                    'Page Title': self.forth_page_title.text(),
                    'Field Name': [],
                    'Selected': [],
                    'Editable': 'Add new button will be added' if self.edit_checkbox.isChecked() else '', 
                    'Location Button': 'Location button will be added' if self.location_checkbox.isChecked() else '',
                    'Delete Button' : 'Delete button will be added' if self.delete_checkbox.isChecked() else ''
                }
                for i in range(self.table_widget.rowCount()):
                    field_name = self.table_widget.item(i, 0).text()
                    selected = self.table_widget.item(i, 1).checkState() == QtCore.Qt.Checked
                    data['Field Name'].append(field_name)
                    data['Selected'].append('Selected' if selected else '')

                df = pd.DataFrame(data)

                df.to_excel(file_path, index=False, engine='openpyxl')
                QtWidgets.QMessageBox.information(self, 'Information', 'Data saved to Excel.')
            except Exception as e:
                QtWidgets.QMessageBox.critical(self, 'Error', f'Error saving data to Excel: {str(e)}')

    def go_back_to_first_page(self):
        self.stacked_widget.setCurrentIndex(4)  
        self.table_widget.clearContents()
        self.table_widget.setRowCount(0)
        self.table_widget.setColumnCount(0)

        self.forth_page_title.clear()

        self.edit_checkbox.setChecked(False)
        self.location_checkbox.setChecked(False)

class UserPrewievPage(QtWidgets.QWidget):
    def __init__(self, stacked_widget):
        super().__init__()
        self.stacked_widget = stacked_widget
        self.num = 0  # Number of fields
        self.setupUi()

    def setupUi(self):
        self.layout = QtWidgets.QVBoxLayout()

        font = QtGui.QFont()
        font.setPointSize(12)

        

        self.table = QtWidgets.QTableWidget()
        self.layout.addWidget(self.table)

        self.setLayout(self.layout)
        self.setGeometry(200, 200, 600, 400)
        self.setWindowTitle('Excel Viewer')
        
        self.load_button = QtWidgets.QPushButton('Open Excel File', self)
        self.load_button.setFont(font)
        self.load_button.clicked.connect(self.load_excel_and_display)
        self.layout.addWidget(self.load_button)

        self.back_button = QtWidgets.QPushButton('Go Back', self)
        self.back_button.setFont(font)
        self.back_button.clicked.connect(self.go_back)
        self.layout.addWidget(self.back_button)

    def load_excel_and_display(self):
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(self, 'Open Excel File', '', 'Excel Files (*.xlsx *.xls)')
        if file_path:
            try:
                df = pd.read_excel(file_path, engine='openpyxl')
                self.display_data(df)
            except Exception as e:
                QtWidgets.QMessageBox.critical(self, 'Error', f'Error loading Excel file: {str(e)}')

    def display_data(self, df):
        font = QtGui.QFont()
        font.setPointSize(15)
        self.table.setFont(font)

        self.table.setRowCount(df.shape[0])
        self.table.setColumnCount(df.shape[1])  
        self.table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        
        
        header_labels = list(df.columns)  
        self.table.setHorizontalHeaderLabels(header_labels)
        
        for i in range(df.shape[0]):
            for j in range(df.shape[1]):
                item = QtWidgets.QTableWidgetItem(str(df.iloc[i, j]))
                self.table.setItem(i, j, item)

            
            checkbox_item = QtWidgets.QTableWidgetItem()
            self.table.setItem(i, df.shape[1], checkbox_item)

        self.table.setColumnWidth(df.shape[1], 70)

    def go_back(self):
        self.stacked_widget.setCurrentIndex(2)
        self.table.setRowCount(0)
        self.table.setColumnCount(0)

class AdminPrewievPage(QtWidgets.QWidget):
    def __init__(self, stacked_widget):
        super().__init__()
        self.stacked_widget = stacked_widget
        self.num = 0  # Number of fields
        self.setupUi()

    def setupUi(self):
        self.layout = QtWidgets.QVBoxLayout()

        font = QtGui.QFont()
        font.setPointSize(1)

        

        self.table = QtWidgets.QTableWidget()
        self.layout.addWidget(self.table)

        self.setLayout(self.layout)
        self.setGeometry(200, 200, 600, 400)
        self.setWindowTitle('Excel Viewer')
        
        self.load_button = QtWidgets.QPushButton('Open Excel File', self)
        self.load_button.setFont(font)
        self.load_button.clicked.connect(self.load_excel_and_display)
        self.layout.addWidget(self.load_button)

        self.back_button = QtWidgets.QPushButton('Go Back', self)
        self.back_button.setFont(font)
        self.back_button.clicked.connect(self.go_back)
        self.layout.addWidget(self.back_button)

    def load_excel_and_display(self):
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(self, 'Open Excel File', '', 'Excel Files (*.xlsx *.xls)')
        if file_path:
            try:
                df = pd.read_excel(file_path, engine='openpyxl')
                self.display_data(df)
            except Exception as e:
                QtWidgets.QMessageBox.critical(self, 'Error', f'Error loading Excel file: {str(e)}')

    def display_data(self, df):
        font = QtGui.QFont()
        font.setPointSize(15)
        self.table.setFont(font)

        self.table.setRowCount(df.shape[0])
        self.table.setColumnCount(df.shape[1])  
        self.table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        
        
        header_labels = list(df.columns)  
        self.table.setHorizontalHeaderLabels(header_labels)
        
        for i in range(df.shape[0]):
            for j in range(df.shape[1]):
                item = QtWidgets.QTableWidgetItem(str(df.iloc[i, j]))
                self.table.setItem(i, j, item)

            
            checkbox_item = QtWidgets.QTableWidgetItem()
            self.table.setItem(i, df.shape[1], checkbox_item)

        self.table.setColumnWidth(df.shape[1], 70)

    def go_back(self):
        self.stacked_widget.setCurrentIndex(11)
        self.table.setRowCount(0)
        self.table.setColumnCount(0)

class UserEditPage(QtWidgets.QWidget):
    
    def __init__(self, stacked_widget):
        super().__init__()
        self.stacked_widget = stacked_widget
        self.num = 0  # Number of fields
        self.setupUi()

    def setupUi(self):

        
        self.layout = QtWidgets.QVBoxLayout()


        self.table = QtWidgets.QTableWidget() #verilerin gozukmesi icin exceldeki gibi
        self.layout.addWidget(self.table)

        font = QtGui.QFont()
        font.setPointSize(15)
        
        self.load_button = QtWidgets.QPushButton('Open Excel File', self)
        self.load_button.setFont(font)
        self.load_button.clicked.connect(self.load_excel_and_display)
        self.layout.addWidget(self.load_button)

        self.save_button = QtWidgets.QPushButton('Save', self)
        self.save_button.setFont(font)
        self.save_button.clicked.connect(self.save_to_excel)
        self.layout.addWidget(self.save_button)

        self.back_button = QtWidgets.QPushButton('Go Back', self)
        self.back_button.setFont(font)
        self.back_button.clicked.connect(self.go_back)
        self.layout.addWidget(self.back_button)

        self.setLayout(self.layout)
        self.setGeometry(200, 200, 600, 400)
        self.setWindowTitle('Excel Viewer')

    def load_excel_and_display(self):
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(self, 'Open Excel File', '', 'Excel Files (*.xlsx *.xls)') 
        if file_path:
            try:
                self.df = pd.read_excel(file_path, engine='openpyxl')
                self.display_data(self.df)
            except Exception as e:
                QtWidgets.QMessageBox.critical(self, 'Error', f'Error loading Excel file: {str(e)}')

    def display_data(self, df):
        font = QtGui.QFont()
        font.setPointSize(15)
        self.table.setFont(font)

        self.table.clear()
        self.table.setRowCount(df.shape[0])
        self.table.setColumnCount(df.shape[1])  # checkbox lar icin (kaldirilabilir)

        header_labels = list(df.columns)
        self.table.setHorizontalHeaderLabels(header_labels)
       
        for i in range(df.shape[0]):
            for j in range(df.shape[1]):
                item = QtWidgets.QTableWidgetItem(str(df.iloc[i, j])) #verilerin gozukmesi icin
                self.table.setItem(i, j, item)

            
            checkbox_item = QtWidgets.QTableWidgetItem()
            checkbox_item.setFlags(QtCore.Qt.ItemIsUserCheckable | QtCore.Qt.ItemIsEnabled)
            checkbox_item.setCheckState(QtCore.Qt.Unchecked)
            self.table.setItem(i, df.shape[1], checkbox_item)

        
        self.table.setEditTriggers(QtWidgets.QAbstractItemView.DoubleClicked | QtWidgets.QAbstractItemView.EditKeyPressed)

       
        self.table.setColumnWidth(df.shape[1], 70)

    def save_to_excel(self):
        file_path, _ = QtWidgets.QFileDialog.getSaveFileName(self, 'Save Excel File', '', 'Excel Files (*.xlsx *.xls)') #kaydetme islemi
        if file_path:
            try:
                
                for i in range(self.table.rowCount()):
                    for j in range(self.table.columnCount() - 1): 
                        self.df.iloc[i, j] = self.table.item(i, j).text()

                
                self.df.to_excel(file_path, index=False)
                QtWidgets.QMessageBox.information(self, 'Success', 'Data saved to Excel.')
            except Exception as e:
                QtWidgets.QMessageBox.critical(self, 'Error', f'Error saving Excel file: {str(e)}')

    def go_back(self):
        self.stacked_widget.setCurrentIndex(2)

class AdminEditPage(QtWidgets.QWidget):
    
    def __init__(self, stacked_widget):
        super().__init__()
        self.stacked_widget = stacked_widget
        self.num = 0  
        self.setupUi()

    def setupUi(self):
        self.layout = QtWidgets.QVBoxLayout()
        
        font = QtGui.QFont()
        font.setPointSize(15)

        self.table = QtWidgets.QTableWidget()
        self.layout.addWidget(self.table)
        
        self.load_button = QtWidgets.QPushButton('Open Excel File', self)
        self.load_button.setFont(font)
        self.load_button.clicked.connect(self.load_excel_and_display)
        self.layout.addWidget(self.load_button)

        self.save_button = QtWidgets.QPushButton('Save', self)
        self.save_button.setFont(font)
        self.save_button.clicked.connect(self.save_to_excel)
        self.layout.addWidget(self.save_button)

        self.back_button = QtWidgets.QPushButton('Go Back', self)
        self.back_button.setFont(font)
        self.back_button.clicked.connect(self.go_back)
        self.layout.addWidget(self.back_button)

        self.setLayout(self.layout)
        self.setGeometry(200, 200, 600, 400)
        self.setWindowTitle('Excel Viewer')

    def load_excel_and_display(self):
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(self, 'Open Excel File', '', 'Excel Files (*.xlsx *.xls)')
        if file_path:
            try:
                self.df = pd.read_excel(file_path, engine='openpyxl')
                self.display_data(self.df)
            except Exception as e:
                QtWidgets.QMessageBox.critical(self, 'Error', f'Error loading Excel file: {str(e)}')

    def display_data(self, df):
        font = QtGui.QFont()
        font.setPointSize(15)
        self.table.setFont(font)

        self.table.clear()
        self.table.setRowCount(df.shape[0])
        self.table.setColumnCount(df.shape[1])  

        header_labels = list(df.columns)
        self.table.setHorizontalHeaderLabels(header_labels)
       
        for i in range(df.shape[0]):
            for j in range(df.shape[1]):
                item = QtWidgets.QTableWidgetItem(str(df.iloc[i, j]))
                self.table.setItem(i, j, item)

            
            checkbox_item = QtWidgets.QTableWidgetItem() #checkbox eklendi
            checkbox_item.setFlags(QtCore.Qt.ItemIsUserCheckable | QtCore.Qt.ItemIsEnabled)
            checkbox_item.setCheckState(QtCore.Qt.Unchecked)
            self.table.setItem(i, df.shape[1], checkbox_item)

       
        self.table.setEditTriggers(QtWidgets.QAbstractItemView.DoubleClicked | QtWidgets.QAbstractItemView.EditKeyPressed)

       
        self.table.setColumnWidth(df.shape[1], 70)

    def save_to_excel(self):
        file_path, _ = QtWidgets.QFileDialog.getSaveFileName(self, 'Save Excel File', '', 'Excel Files (*.xlsx *.xls)')
        if file_path:
            try:
                
                for i in range(self.table.rowCount()): #verilerin kaydedilmesi
                    for j in range(self.table.columnCount() - 1):   
                        self.df.iloc[i, j] = self.table.item(i, j).text() #olusturulan table daki veriler kaydediliyor teker teker gezinerek

                
                self.df.to_excel(file_path, index=False)
                QtWidgets.QMessageBox.information(self, 'Success', 'Data saved to Excel.')
            except Exception as e:
                QtWidgets.QMessageBox.critical(self, 'Error', f'Error saving Excel file: {str(e)}')

    def go_back(self):
        self.stacked_widget.setCurrentIndex(11)



class RequirementPage(QWidget):
    def __init__(self, stacked_widget):
        super().__init__()
        self.stacked_widget = stacked_widget
        self.current_sentence_index = 0
        self.current_sentence_layout = None
        self.current_row_layout = None
        self.setupUi()
        #self.checkedChanged = pyqtSignal(bool)

    def setupUi(self):
        self.setWindowTitle('Dynamic QLineEdit Example')
        self.setGeometry(100, 100, 600, 400)

        self.main_layout = QVBoxLayout()
        self.setLayout(self.main_layout)

        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)

        self.scroll_content = QWidget()
        self.scroll_layout = QVBoxLayout()
        self.scroll_content.setLayout(self.scroll_layout)

        self.scroll_area.setWidget(self.scroll_content)
        self.main_layout.addWidget(self.scroll_area)

        dummy_line_edit = QLineEdit()  
        self.line_edit_width = dummy_line_edit.sizeHint().width()
        self.line_edit_height = dummy_line_edit.sizeHint().height()

        self.title_of_name = QLabel("Name of template:")
        self.title_of_name.setFont(QtGui.QFont("Arial", 12))
        self.scroll_layout.addWidget(self.title_of_name)

        self.name_of_template = QLineEdit()
        
        self.name_of_template.setFont(QtGui.QFont("Arial", 12))
        self.scroll_layout.addWidget(self.name_of_template)

        self.new_sentence_button = QPushButton('New sentence')
        self.new_sentence_button.setFont(QtGui.QFont("Arial", 12))
        self.new_sentence_button.clicked.connect(self.new_sentence_clicked)
        
        self.main_layout.addWidget(self.new_sentence_button)

        self.add_button = QPushButton('Add')
        self.add_button.setFont(QtGui.QFont("Arial", 12))
        self.add_button.clicked.connect(self.add_button_clicked)
        self.main_layout.addWidget(self.add_button)

        self.secim_combobox = QComboBox(self)
        self.secim_combobox.addItem("")
        self.secim_combobox.addItem("phrase")
        self.secim_combobox.addItem("parameter")
        self.secim_combobox.setVisible(False)
        self.main_layout.addWidget(self.secim_combobox)
        self.secim_combobox.setCurrentText("")
        self.secim_combobox.setFont(QtGui.QFont("Arial", 12))

        self.concatenate_button = QPushButton('Save')
        self.concatenate_button.setFont(QtGui.QFont("Arial", 12))
        self.concatenate_button.clicked.connect(self.concatenate_texts)
        self.main_layout.addWidget(self.concatenate_button)

        self.go_back_button = QPushButton()
        self.go_back_button.setText("Go Back")
        self.go_back_button.setFont(QtGui.QFont("Arial", 12))
        self.go_back_button.clicked.connect(self.go_back_to_admindec_page)
        self.main_layout.addWidget(self.go_back_button)

        self.help_button = QPushButton()
        self.help_button.setText("Help")
        self.help_button.setFont(QtGui.QFont("Arial", 12))
        self.help_button.clicked.connect(self.show_help_message)
        self.main_layout.addWidget(self.help_button)

        

        self.line_edits_list = []  
        self.sentence_widgets = []  
        self.sentence_array = []  

        self.sentences_data = {}  
        self.parameter_data = {} 
        
        self.optional_info = []

    def add_line_edit(self, is_parameter=False):
        if self.current_sentence_layout is None:
            QMessageBox.warning(self, "Error", "Please create a new sentence first.")
            return

        if self.secim_combobox.currentText() == "":
            QMessageBox.warning(self, "Error", "Please choose an option.")
            return

        if self.current_row_layout is None or self.current_row_layout.count() >= 10: #10 dan fazla oldugunda asagi inecek
            self.current_row_layout = QHBoxLayout() 
            self.current_sentence_layout.addLayout(self.current_row_layout)
            self.current_row_layout.setAlignment(Qt.AlignLeft) #soldan baslamasi icin

        new_line_edit = QLineEdit() # yeni line edit olusturuldu
        new_line_edit.setFont(QtGui.QFont("Arial", 12)) 
        new_line_edit.setFixedSize(self.line_edit_width, self.line_edit_height)
        self.current_row_layout.addWidget(new_line_edit) 
        self.line_edits_list.append(new_line_edit)

        if is_parameter: #eger paremeter ise
            font_italic = new_line_edit.font()
            font_italic.setUnderline(True)
            font_italic.setItalic(True)
            new_line_edit.setFont(font_italic)
            new_line_edit.setStyleSheet("QLineEdit"
                                "{"
                                "background : lightblue;"
                                "}") 

        self.current_row_layout.addWidget(new_line_edit)
        self.line_edits_list.append(new_line_edit)
        self.sentence_widgets[-1].append(new_line_edit)

        #self.sentence_widgets[-1].append(new_line_edit)

        if is_parameter: #parameter ise
            if self.current_sentence_index not in self.parameter_data: #eger parameter data yoksa
                self.parameter_data[self.current_sentence_index] = [] #bos liste olusturuldu
               

            self.parameter_data[self.current_sentence_index].append(new_line_edit)
            

        self.scroll_area.ensureWidgetVisible(new_line_edit) #eklenen line edit in gozukmesi icin

    def add_button_clicked(self): 
        
        if self.secim_combobox.currentText() == "parameter": #parametet secilirse
            self.add_line_edit(is_parameter=True)
            
            
        else:
            self.add_line_edit()

    def new_sentence_clicked(self):

        self.secim_combobox.show()

        self.current_sentence_index += 1

        if self.current_sentence_index == 1: # 1st end 4rd yazabilmesi icin sentence lara
            heading_text = "1st"
        elif self.current_sentence_index == 2:
            heading_text = "2nd"
        elif self.current_sentence_index == 3:
            heading_text = "3rd"
        else:
            heading_text = f"{self.current_sentence_index}th"

        sentence_container_layout = QVBoxLayout()

        new_title_label = QLabel(f'{heading_text} Sentence:')
        new_title_label.setFont(QtGui.QFont("Arial", 12))
        self.sentence_array.append(new_title_label)
        sentence_container_layout.addWidget(new_title_label)

        self.optional_checkbox = QCheckBox("Optional")
        self.optional_checkbox.setFont(QtGui.QFont("Arial", 12))
        sentence_container_layout.addWidget(self.optional_checkbox)
        self.optional_info.append(self.optional_checkbox)
       

        if self.optional_checkbox.isChecked():
            self.checkbox_checked = True
            
        self.current_sentence_layout = QVBoxLayout()
        sentence_container_layout.addLayout(self.current_sentence_layout)

        self.scroll_layout.addLayout(sentence_container_layout)

       
        self.sentence_widgets.append([])
        
        self.sentences_data[heading_text] = ""

        self.current_row_layout = None

    def concatenate_texts(self):
        for index, sentence_list in enumerate(self.sentence_widgets):
            heading_text = f"{index + 1}th" if index > 2 else ["1st", "2nd", "3rd"][index]
            combined_text = ' '.join(edit.text() for edit in sentence_list)
            self.sentences_data[heading_text] = combined_text.strip()

        for key, value in self.sentences_data.items():
            print(f"{key}: {value}")

        self.save_to_excel()

    def checkbox_checked(self):
        checked = True  
        self.checkedChanged.emit(checked) 


        
    def save_to_excel(self):
        
        current_parameters = [param.text() for params_list in self.parameter_data.values() for param in params_list]

        
        existing_parameters = []
        file_path = 'requirement_templates.xlsx'
        if os.path.isfile(file_path):
            existing_data = pd.read_excel(file_path, engine='openpyxl')
            if 'Parameters' in existing_data.columns:
                for param_str in existing_data['Parameters']:
                    if pd.notna(param_str):
                        existing_parameters.extend(param_str.split(', '))

        
        duplicates = set(current_parameters) & set(existing_parameters)
        if duplicates:
            duplicate_params = ', '.join(duplicates)
            reply = QMessageBox.question(self, 'Duplicate Parameters',
                                        f'The parameters already exist: {duplicate_params}\nDo you want to save the data?',
                                        QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

            if reply == QMessageBox.No:
                return

       
        data = []
        for index, sentence_list in enumerate(self.sentence_widgets):
            heading_text = f"{index + 1}th" if index > 2 else ["1st", "2nd", "3rd"][index] #1st 2nd 3rd yazilmasi icin
            combined_text = ' '.join(edit.text() for edit in sentence_list)
            parameters = [param.text() for param in self.parameter_data.get(index + 1, [])]
            optional = self.optional_info[index].isChecked() if index < len(self.optional_info) else False

            data.append({
                'Sentence': heading_text,
                'Text': combined_text,
                'Parameters': ', '.join(parameters),
                'Template Name': self.name_of_template.text(),
                'Optional': optional
            })

        try:
            if not os.path.isfile(file_path): #dosya yoksa
                df = pd.DataFrame(data)
                df.to_excel(file_path, index=False, engine='openpyxl')
            else:
                existing_data = pd.read_excel(file_path, engine='openpyxl') #var olan dosyayi okuyoruz
                df = pd.concat([existing_data, pd.DataFrame(data)], ignore_index=True)
                df.to_excel(file_path, index=False, engine='openpyxl')

            workbook = load_workbook(file_path)
            worksheet = workbook['Sheet1']

            italic_underline_font = Font(italic=True, underline='single')

            for index, row in df.iterrows(): #exceldeki verilerin duzenlenmesi
                if 'Parameters' in row: 
                    cell = worksheet.cell(row=index + 2, column=df.columns.get_loc('Parameters') + 1)  
                    cell.font = italic_underline_font

            for col in worksheet.columns: #sutunlarin genisligi (cok onemli degil)
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                            pass
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[column].width = adjusted_width

            workbook.save(file_path)

            QMessageBox.information(self, "Success", "Data saved successfully.")
        except PermissionError:
            QMessageBox.warning(self, "Error", "Permission denied: Unable to save file. Please check file permissions.")
        except Exception as e:
            QMessageBox.warning(self, "Error", f"An error occurred while saving data:\n{str(e)}")


                
            
    def show_help_message(self): #help butonunun mesaji
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setText("How to use this page?")
        msg.setInformativeText("First, use 'New Sentence' button to create a new sentence. Then choose a type for your content from the combobox. Finally, add content into your sentence using the 'Add' button.")
        msg.setWindowTitle("Help")
        msg.exec_()

    def go_back_to_admindec_page(self):
        #print("Going back to Admin Decision Page")
        self.stacked_widget.setCurrentIndex(11)
        self.reset_ui()

    def reset_ui(self):
            # Clear all QLineEdit widgets
        for line_edit_list in self.sentence_widgets:
            for line_edit in line_edit_list:
                line_edit.deleteLater()

      
        
        
        self.current_sentence_index = 0
        self.current_sentence_layout = None
        self.current_row_layout = None
        
        for label in self.sentence_array:
            label.deleteLater()
        
        self.sentence_array = []

        self.name_of_template.clear()
        self.secim_combobox.hide()

        self.line_edits_list = []
        self.sentence_widgets = []
        self.sentences_data = {}
        self.parameter_data = {}

        self.optional_checkbox.deleteLater()
        

class RequirementEditPage(QWidget):
    def __init__(self, stacked_widget):
        super().__init__()
        self.stacked_widget = stacked_widget
        self.parameter_list = []
        self.phrase_list = []
        self.setupUi()

    def setupUi(self):
        self.tab_part = QTabWidget()

        layout = QVBoxLayout(self)
        layout.addWidget(self.tab_part)
        self.setLayout(layout)

        self.load_tab = QWidget()
        self.edit_tab = QWidget()

        self.tab_part.addTab(self.load_tab, "Load")
        self.tab_part.addTab(self.edit_tab, "Edit")    

        self.setup_load_tab()

        tab_settings = self.tab_part.tabBar() #tablarin arasindaki bosluklari ayarlar
        font = QFont("Arial", 12)
        tab_settings.setFont(font)

        

    def setup_load_tab(self):
        layout = QVBoxLayout(self.load_tab)

        #gerekgli ui urunlerinin olusmasi
        self.search_bar = QLineEdit(self) 
        self.search_bar.setPlaceholderText("Search 🔍︎")
        layout.addWidget(self.search_bar)
        self.search_bar.setFont(QFont("Arial", 12))
        
       
        self.template_name_list = QListWidget()
        layout.addWidget(self.template_name_list)
        self.template_name_list.setFont(QFont("Arial", 12))
        
        
        button_layout = QHBoxLayout()
        layout.addLayout(button_layout)
        
        
        self.load_button = QPushButton("Load")
        self.load_button.clicked.connect(self.load_template)
        button_layout.addWidget(self.load_button)
        self.load_button.setFont(QFont("Arial", 15))
        self.load_button.setFixedSize(100, 50)

        
        self.refresh_button = QPushButton("Refresh")
        self.refresh_button.clicked.connect(self.refresh_data)
        button_layout.addWidget(self.refresh_button)
        self.refresh_button.setFont(QFont("Arial", 15))
        self.refresh_button.setFixedSize(100, 50)

       
        self.go_back_button = QPushButton("Go Back")
        button_layout.addWidget(self.go_back_button)
        self.go_back_button.setFont(QFont("Arial", 15))
        self.go_back_button.setFixedSize(100, 50)
        self.go_back_button.clicked.connect(self.go_back)

       
        self.data_display = QLabel() 
        layout.addWidget(self.data_display)

        
        self.load_template_names()
        self.current_data = None
        self.line_edits = []

        
        self.search_bar.textChanged.connect(self.filter_list)
        

    def load_template_names(self):
        try:
            df = pd.read_excel('requirement_templates.xlsx', engine='openpyxl') #varolan dosyayi okuyoruz
            template_names = df['Template Name'].unique() #unique olanlari aliyoruz 
            self.template_name_list.addItems(template_names)
            self.items = template_names  # arama cubugu icin gerekli

        except Exception as e:
            QMessageBox.warning(self, "Error", f"An error occurred:\n{str(e)}")

    def filter_list(self, text):
        filtered_items = [item for item in self.items if item.lower().startswith(text.lower())]
        self.template_name_list.clear()
        self.template_name_list.addItems(filtered_items)

    def load_template(self):
        selected_items = self.template_name_list.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Error", "Please select a template name.")
            return
        
        template_name = selected_items[0].text()

        try:
            df = pd.read_excel('requirement_templates.xlsx', engine='openpyxl')
            self.current_data = df[df['Template Name'] == template_name] 

            self.display_data()
            self.edit_tab_func(self.current_data)
        except Exception as e:
            QMessageBox.warning(self, "Error", f"An error occurred:\n{str(e)}")

    def display_data(self):
        colum_names = ["Template Name", "Sentence", "Text", "Parameters", "Optional"] #bu column lar var excelde onlar okundu
        if not all(col in self.current_data.columns for col in colum_names):
            QMessageBox.warning(self, "Error", "Couldn't find required columns.")
            return
        
        wanted_data = self.current_data
        
        if wanted_data.empty:
            QMessageBox.warning(self, "Error", "Couldn't find any data.")
            return
        else:
            self.data_display.setText(f"{self.get_display_text(wanted_data)}") 
            font = self.data_display.font()
            font.setPointSize(14)  
            self.data_display.setFont(font)
            

    def get_display_text(self, data):
        text = ""
        
    
        for index, row in data.iterrows(): #verileri alir
            text += f"Sentence: {row['Sentence']}\n {row['Text']}\n " f"Optional: {row['Optional']}\n\n" 
        return text

    def edit_tab_func(self, data):
        current_layout = self.edit_tab.layout() 
        if current_layout: 
            QWidget().setLayout(current_layout)   
        
        layout = QVBoxLayout(self.edit_tab)
        self.line_edits = []  # line_Editleri tutmak icin liste
        
        for index, row in data.iterrows():
            sentence_label = QLabel(row['Sentence']) 
            layout.addWidget(sentence_label) 
            sentence_label.setFont(QtGui.QFont("Arial", 13)) 
            
            self.text_fields = str(row['Text']).split(" ")
            parameters = [parameter.strip() for parameter in str(row["Parameters"]).split(",")]
            
            row_edit = []
            parameter_edit = []
            row_layout = QHBoxLayout()
            
            for text in self.text_fields:    
                self.line_edit = QLineEdit()
                self.line_edit.setText(text)
                font = self.line_edit.font()
                font.setPointSize(13)
                self.line_edit.setFont(font)
                
                if text in parameters:
                    font.setItalic(True)
                    font.setUnderline(True)
                    self.line_edit.setFont(font)
                    self.line_edit.setStyleSheet("background-color: lightblue;")
                    parameter_edit.append(self.line_edit)
                
                row_layout.addWidget(self.line_edit)
                row_edit.append(self.line_edit)
            
            layout.addLayout(row_layout)
            self.line_edits.append((index, row_edit, parameter_edit))
            
            
            optional_label = QLabel("Optional:")
            layout.addWidget(optional_label)
            optional_option = QLineEdit(f"{row.get('Optional', '')}")
            layout.addWidget(optional_option)
        
        button_layout = QHBoxLayout()
        layout.addLayout(button_layout)
        
        self.save_button = QPushButton("Save")
        self.save_button.clicked.connect(self.save_changes)
        button_layout.addWidget(self.save_button)
        self.save_button.setFont(QtGui.QFont("Arial", 15))

    def save_changes(self):
        if self.current_data is None:
            QMessageBox.warning(self, "Error", "No data to save.")
            return
        
        try:
            df = pd.read_excel('requirement_templates.xlsx', engine='openpyxl')
            for index, row_edit, parameter_edit in self.line_edits:
                new_text = ' '.join([line_edit.text() for line_edit in row_edit])
                new_parameters = ', '.join([line_edit.text() for line_edit in parameter_edit])

                parameter_list = new_parameters.split(", ")
                right_parameters = [parameter for parameter in parameter_list if parameter in new_text] #parametrelerin textte olup olmadigini kontrol eder
                final_parameters = ', '.join(right_parameters) #dogru parametreleri alir

                df.at[index, 'Text'] = new_text
                df.at[index, 'Parameters'] = final_parameters

            df.to_excel('requirement_templates.xlsx', index=False, engine='openpyxl')
            QMessageBox.information(self, "Success", "Data saved successfully.")

        except Exception as e:
            QMessageBox.warning(self, "Error", f"An error occurred:\n{str(e)}")

    def refresh_data(self):
        self.load_template()
        #self.data_display.setText("")
        #self.current_data = None
        #self.line_edits = []
    def go_back(self):
        self.stacked_widget.setCurrentIndex(11)
        self.reset_ui()
    
    def reset_ui(self):
       
        self.data_display.setText("")
        self.current_data = None
        self.wanted_data = None
        self.line_edits = []
        self.line_edits.clear()
        self.current_data = None
        current_layout = self.edit_tab.layout()
        if current_layout:
            QWidget().setLayout(current_layout)
        

   
def create_excel_file():
    file_path = 'requirement_templates.xlsx'
    if not os.path.isfile(file_path):
        data = {
            'Sentence': [],
            'Text': [],
            'Parameters': [],
            'Template Name': []
        }
        df = pd.DataFrame(data)
        df.to_excel(file_path, index=False, engine='openpyxl')
        print(f"{file_path} has been created with the required columns.")


class MainWindow(QStackedWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def get_baslangic_combobox_value(self):
        return self.baslangic_page.get_combobox_value()

    
    
    def initUI(self): #pencereleri olusturma
        self.opening_page = OpeningPage(self) #1
        self.user_signin_page = UserSignInPage(self) #2
        self.user_decision_page = UserDecisionPage(self) #3
        self.baslangic_page = BaslangicPenceresi(self) #4 
        self.first_page = GirisPencere(self)  #5
        self.second_page = SecondPencere(self) #6
        self.third_page = ThirdPencere(self) #7
        self.forth_page = ForthPencere(self) #8
        self.user_preview_page = UserPrewievPage(self) #9
        self.user_edit_page = UserEditPage(self) #↨10
        self.admin_signin_page = AdminSignInPage(self) #11
        self.admin_decision_page = AdminDecisionPage(self) #12
        self.requirement_page = RequirementPage(self) #13
        self.reqed_page = RequirementEditPage(self) #14
        self.admin_edit_page = AdminEditPage(self) #15
        self.admin_preview_page = AdminPrewievPage(self) #16 bu sayilarin bir eksigi indexleri, cagirabilmek icin bi eksigi yazilir
        
        
        self.addWidget(self.opening_page)
        self.addWidget(self.user_signin_page)
        self.addWidget(self.user_decision_page)
        self.addWidget(self.baslangic_page) 
        self.addWidget(self.first_page)
        self.addWidget(self.second_page)
        self.addWidget(self.third_page)
        self.addWidget(self.forth_page)
        self.addWidget(self.user_preview_page)
        self.addWidget(self.user_edit_page)
        self.addWidget(self.admin_signin_page)
        self.addWidget(self.admin_decision_page)
        self.addWidget(self.requirement_page)
        self.addWidget(self.reqed_page)
        self.addWidget(self.admin_edit_page)
        self.addWidget(self.admin_preview_page)

        self.setCurrentIndex(0)

        #self.opening_page.user_pushButton.clicked.connect(self.show_user_signin_page)
        #self.opening_page.admin_pushButton_2.clicked.connect(self.show_admin_signin_page)

        #self.signin_page.pushButton.clicked.connect(self.signin_page.signin_clicked)

    
    def show_user_signin_page(self):
        self.setCurrentIndex(1)
        self.resize(302, 311) 

    def show_admin_signin_page(self):
        self.setCurrentIndex(10)
        self.resize(302, 311)
    def show_user_decision_page(self):
        self.setCurrentIndex(2)
        
    def show_admin_decision_page(self):
        self.setCurrentIndex(11)
        #self.resize(558, 308)

if __name__ == '__main__': #programın calistirilmasi
    create_excel_file()
    app = QApplication(sys.argv)
    main_app = MainWindow()
    main_app.setWindowTitle('Project')
    main_app.setGeometry(300, 300, 400, 300)
    main_app.show()
    sys.exit(app.exec_()) #programi kapatmak icin

    

